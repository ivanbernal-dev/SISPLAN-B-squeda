#!/usr/bin/env python3
"""
Genera archivos Excel de prueba para los templates REALES del sistema UBPD.

Uso:
  python test-data/excel/generar_excel_prueba.py [BASE_URL] [USUARIO] [PASSWORD]

  Defaults:
    BASE_URL = http://localhost/api
    USUARIO  = admin
    PASSWORD = Admin@UBPD2024!

Genera un Excel por template activo en la carpeta test-data/excel/
Cada archivo está listo para subir en: Dependencia → Nuevo Registro → Cargar Excel
"""

import sys
import io
from pathlib import Path

try:
    import requests
except ImportError:
    print("ERROR: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL = sys.argv[1] if len(sys.argv) > 1 else "http://localhost/api"
USUARIO  = sys.argv[2] if len(sys.argv) > 2 else "admin"
PASSWORD = sys.argv[3] if len(sys.argv) > 3 else "Admin@UBPD2024!"
OUT_DIR  = Path(__file__).parent

# ── Estilos ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1A3C5E")
TYPE_FILL    = PatternFill("solid", fgColor="E3F2FD")
REQ_FILL     = PatternFill("solid", fgColor="E8F5E9")
OPT_FILL     = PatternFill("solid", fgColor="FFF8E1")
RO_FILL      = PatternFill("solid", fgColor="EEEEEE")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
TYPE_FONT    = Font(color="1A3C5E", italic=True, size=9, name="Calibri")
DATA_FONT    = Font(size=10, name="Calibri")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def ftype(f): return str(f.get("type", f.get("tipo", "text"))).lower()
def fname(f): return str(f.get("name", "")).strip()
def flabel(f): return str(f.get("label", f.get("name", ""))).strip()
def fopts(f): return f.get("options") or f.get("opciones") or []
def freq(f): return bool(f.get("required", f.get("requerido", False)))
def fro(f):  return bool(f.get("readonly", f.get("bloqueado", False)))


def get_type_hint(f) -> str:
    ft = ftype(f)
    if fro(f):    return "Solo lectura"
    if ft == "number": return ("* " if freq(f) else "") + "Número"
    if ft == "date":   return ("* " if freq(f) else "") + "Fecha (AAAA-MM-DD)"
    if ft == "select":
        opts = fopts(f)
        s = " / ".join(str(o if not isinstance(o, dict) else o.get("value", o)) for o in opts[:3])
        return ("* " if freq(f) else "") + (f"Opciones: {s}" if s else "Seleccionar")
    if ft in ("textarea", "text"):
        return ("* " if freq(f) else "") + ("Texto largo" if ft == "textarea" else "Texto")
    return "Texto"


def get_sample(f, row_idx: int) -> object:
    ft = ftype(f)
    fn = fname(f)
    if fro(f):
        return str(f.get("default") or "")
    if ft == "date":
        months = ["2026-01-31", "2026-02-28", "2026-03-31", "2026-04-30",
                  "2026-05-31", "2026-06-30", "2026-07-31"]
        return months[row_idx % len(months)]
    if ft == "number":
        vals = [0.75, 0.5, 1.0, 0.25, 0.8, 0.6, 0.9]
        return vals[row_idx % len(vals)]
    if ft == "select":
        opts = fopts(f)
        if opts:
            o = opts[row_idx % len(opts)]
            return o if not isinstance(o, dict) else o.get("value", str(o))
        return ""
    if fn == "informe_cualitativo":
        examples = [
            "Durante el mes se avanzó en la revisión de documentos y coordinación con equipos territoriales.",
            "Se realizaron reuniones de seguimiento y se ajustaron los cronogramas de trabajo.",
            "El equipo completó las actividades planificadas con un avance del 80% sobre la meta.",
            "Se identificaron obstáculos logísticos que fueron resueltos en la segunda semana del mes.",
            "Se entregaron los productos acordados y se inició la siguiente fase de implementación.",
            "Avance sostenido en las actividades clave; pendiente validación por parte de la dirección.",
            "El indicador muestra tendencia positiva respecto al trimestre anterior.",
        ]
        return examples[row_idx % len(examples)]
    if "variable" in fn and "cuantitativo" not in fn:
        labels = [
            "Documentos publicados en la plataforma institucional",
            "Reuniones técnicas realizadas con equipos regionales",
            "Informes de avance entregados a la dirección",
            "Actividades de campo ejecutadas en territorio",
            "Espacios de socialización con comunidades",
            "Jornadas de capacitación realizadas",
            "Protocolos actualizados y validados",
        ]
        return labels[row_idx % len(labels)]
    if "meta" in fn:
        return "100% de los productos entregados antes del cierre del segundo semestre."
    if "entregable" in fn:
        idx = fn[-1] if fn[-1].isdigit() else "1"
        items = [
            "Informe de avance mensual consolidado",
            "Acta de reunión de seguimiento",
            "Documento de sistematización de resultados",
            "Reporte de indicadores de gestión",
            "Matriz de seguimiento actualizada",
        ]
        return items[row_idx % len(items)]
    if "eje" in fn or "actividade" in fn:
        return "Fortalecimiento de capacidades técnicas y articulación interinstitucional."
    return str(f.get("default") or "Ejemplo")


def make_excel(template: dict, n_rows: int = 5) -> str:
    """Genera Excel de prueba para un template. Retorna el nombre del archivo."""
    cfg = template.get("configuracion_campos") or {}
    all_fields = cfg.get("fields") or cfg.get("campos") or []
    # Excluir computed y archivos
    fields = [f for f in all_fields if ftype(f) not in ("computed", "archivos", "file")]

    wb = Workbook()
    ws = wb.active
    ws.title = (template.get("codigo") or template["nombre"])[:31]

    # Encabezados
    ws.append([flabel(f) for f in fields])
    # Tipo/guía
    ws.append([get_type_hint(f) for f in fields])
    # Filas de datos
    for i in range(n_rows):
        ws.append([get_sample(f, i) for f in fields])

    # Estilos
    for ci, f in enumerate(fields, start=1):
        letter = get_column_letter(ci)
        is_ro  = fro(f)
        is_req = freq(f)
        ws.column_dimensions[letter].width = 32 if ftype(f) in ("textarea", "select") else 26

        # Header
        c = ws.cell(row=1, column=ci)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER

        # Type hint
        c2 = ws.cell(row=2, column=ci)
        c2.font = TYPE_FONT; c2.fill = TYPE_FILL; c2.alignment = LEFT; c2.border = BORDER

        # Data rows
        for ri in range(n_rows):
            c3 = ws.cell(row=3 + ri, column=ci)
            c3.font = DATA_FONT
            c3.fill = RO_FILL if is_ro else (REQ_FILL if is_req else OPT_FILL)
            c3.alignment = LEFT; c3.border = BORDER

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 110

    # Hoja de referencia
    ws2 = wb.create_sheet("Referencia")
    ws2.append(["Campo", "Nombre interno", "Tipo", "Requerido", "Solo lectura", "Opciones"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="37474F")
        cell.alignment = CENTER
    for f in fields:
        opts = ", ".join(str(o if not isinstance(o, dict) else o.get("value", o)) for o in fopts(f))
        ws2.append([flabel(f), fname(f), ftype(f),
                    "Sí" if freq(f) else "No",
                    "Sí" if fro(f) else "No",
                    opts])
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 35

    # Nombre de archivo seguro
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in template.get("codigo", template["nombre"]))
    filename = f"{safe}.xlsx"
    wb.save(OUT_DIR / filename)
    return filename


def main():
    print(f"\nConectando a {BASE_URL} como '{USUARIO}'...\n")
    session = requests.Session()

    # Login
    r = session.post(f"{BASE_URL}/auth/login",
                     json={"username": USUARIO, "password": PASSWORD}, timeout=10)
    if r.status_code != 200:
        print(f"ERROR: Login falló ({r.status_code}): {r.text[:200]}")
        sys.exit(1)
    token = r.json().get("access_token")
    session.headers.update({"Authorization": f"Bearer {token}"})

    # Obtener templates
    r = session.get(f"{BASE_URL}/templates", timeout=15)
    if r.status_code != 200:
        print(f"ERROR: No se pudieron obtener templates ({r.status_code})")
        sys.exit(1)
    templates = r.json()
    print(f"Templates encontrados: {len(templates)}\n")

    generated = []
    for t in templates:
        try:
            fn = make_excel(t, n_rows=5)
            generated.append((t["nombre"], fn))
            print(f"  ✓ {fn}")
            print(f"    Template: {t['nombre']}")
        except Exception as e:
            print(f"  ✗ Error en '{t['nombre']}': {e}")

    print(f"\n✓ {len(generated)} archivo(s) generados en: {OUT_DIR}\n")
    print("Cómo usarlos:")
    print("  1. Inicia sesión como usuario de dependencia")
    print("  2. Ve a 'Nuevo Registro' y selecciona el template correspondiente")
    print("  3. Cambia a la pestaña 'Cargar Excel'")
    print("  4. Sube el archivo .xlsx correspondiente al template")
    print("  5. El validador verá todos los registros del lote para aprobar/rechazar\n")


if __name__ == "__main__":
    main()
