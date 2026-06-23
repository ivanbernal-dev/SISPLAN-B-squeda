"""
Genera un Excel de ejemplo para CADA template PAI 2026, con datos sintéticos
realistas: 8 filas por template (2 actividades × 4 trimestres) con
% Avance Proyectado y Alcanzado verosímiles.

Salida: scripts/pai_2026/examples/{codigo_template}_ejemplo.xlsx

Estos archivos pueden cargarse desde una dependencia para probar el flujo
end-to-end (upload → validador aprueba → pipeline calcula → estadísticas).
"""
from __future__ import annotations

import json
import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).parent
DATA_DIR = ROOT / "data"
OUT_DIR  = ROOT / "examples"
OUT_DIR.mkdir(exist_ok=True)

TEMPLATES_JSON = DATA_DIR / "pai_templates.json"

# Reproducibilidad
random.seed(42)


def _round(x, n=2):
    return round(float(x), n)


def synth_value(field: dict, trim_idx: int, act_idx: int, codigo: str) -> object:
    """Genera un valor sintético para un campo dado el contexto."""
    name = field.get("name", "")
    ftype = field.get("type", "text")
    default = field.get("default")

    # Constantes del producto → siempre el default extraído del Excel original
    if field.get("readonly") and default is not None and not field.get("auto_calculate"):
        return default

    # Campos auto_calculate o validator_only → vacíos (los recalcula sistema / valida OAP)
    if field.get("auto_calculate") or field.get("validator_only"):
        return None

    # Periodo de reporte (selector trimestre)
    if name == "periodo_reporte":
        return f"TRIMESTRE {trim_idx}"

    # Ejes (varia por actividad)
    if name == "eje":
        return f"{act_idx}. Eje de ejemplo para actividad {act_idx}"

    # Pesos numéricos
    if name == "peso_actividad":
        return 0.5 if act_idx == 1 else 0.5
    if name == "peso_trimestre":
        return _round(0.20 + random.uniform(-0.05, 0.05))

    # Actividades / indicador / metas → texto descriptivo
    if name == "actividad_clave":
        return f"{act_idx}.{trim_idx} Actividad clave de ejemplo — {codigo} T{trim_idx}"
    if name == "indicador":
        return f"• Indicador de ejemplo para la actividad {act_idx} en T{trim_idx}"
    if name == "meta_anual":
        return "100% de la meta anual"
    if name == "entregable_total":
        return "Entregable consolidado del producto"
    if name == "entregable_trimestre":
        return f"Entregable de T{trim_idx} actividad {act_idx}"

    # % Avance proyectado y alcanzado — valores realistas
    if name == "pct_avance_proyectado":
        # Proyección acumulada por trimestre: 0.25, 0.5, 0.75, 1.0
        return _round(0.25 * trim_idx, 2)
    if name == "pct_avance_alcanzado":
        # Alcance acumulado: cerca del proyectado, con ruido (alguna actividad cumple, otra no)
        base = 0.25 * trim_idx
        # Algunas actividades cumplen >90%, otras parcial, otras no cumplen
        bias = {1: 0.92, 2: 0.78, 3: 0.65}.get(act_idx, 0.85)
        noise = random.uniform(-0.05, 0.05)
        return _round(min(1.0, max(0.0, base * bias + noise)), 2)

    # Comentarios — solo el primero
    if name == "avances_logros":
        return f"Avances reportados en T{trim_idx} — actividad {act_idx}: cumplimiento parcial."
    if name == "retrasos_dificultades":
        return "Sin retrasos significativos."
    if name == "comentarios_extra":
        return ""

    # Default
    if ftype == "number":
        return 0
    return ""


def build_excel(template: dict) -> Path:
    codigo = template["codigo"]
    fields = template["configuracion_campos"]["fields"]
    # Columnas a exportar = lo mismo que el endpoint admin/templates/{id}/excel-example:
    #   excluir validator_only y auto_calculate
    cols = [f for f in fields if not f.get("validator_only") and not f.get("auto_calculate")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = codigo[:31]

    # Estilos
    header_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    type_fill   = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    type_font   = Font(color="1A3C5E", italic=True, size=9, name="Calibri")
    border = Border(left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC'))

    # Fila 1: encabezados
    headers = [f["label"] for f in cols]
    ws.append(headers)

    # Fila 2: tipo / guía
    hints = []
    for f in cols:
        ft = f["type"]
        ro = f.get("readonly")
        req = f.get("required")
        if ro:                  h = "Solo lectura"
        elif ft == "number":    h = "Número"
        elif ft == "date":      h = "Fecha (AAAA-MM-DD)"
        elif ft == "select":    h = "Opciones: " + " / ".join(f.get("options") or [])
        elif ft == "textarea":  h = "Texto largo"
        else:                   h = "Texto"
        if not ro:
            h = ("* " if req else "") + h
        hints.append(h)
    ws.append(hints)

    # Filas 3+: 2 actividades × 4 trimestres = 8 filas
    for trim in range(1, 5):
        for act in (1, 2):
            row = [synth_value(f, trim, act, codigo) for f in cols]
            ws.append(row)

    # Estilos cabecera
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).fill = header_fill
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=1, column=c).border = border
        ws.cell(row=2, column=c).font = type_font
        ws.cell(row=2, column=c).fill = type_fill
        ws.cell(row=2, column=c).border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 28
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"

    out = OUT_DIR / f"{codigo}_ejemplo.xlsx"
    wb.save(out)
    return out


def main():
    payload = json.loads(TEMPLATES_JSON.read_text(encoding="utf-8"))
    print(f"Generando {len(payload)} Excels de ejemplo en {OUT_DIR}/ ...")
    for t in payload:
        out = build_excel(t)
        print(f"  ✓ {out.name}  ({out.stat().st_size // 1024} KB)")
    print(f"\nListo. {len(list(OUT_DIR.glob('*.xlsx')))} archivos en {OUT_DIR}")


if __name__ == "__main__":
    main()
