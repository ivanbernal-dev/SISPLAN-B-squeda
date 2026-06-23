"""
Limpia y reconstruye la jerarquía de indicadores + templates a partir del Excel
PAI DEFINITIVO. Convención:
  - Indicador Nivel 1 = Línea Estratégica (L1, L2, ..., L6)
  - Indicador Nivel 2 = Producto del PAI (cada hoja del Excel)
  - Template          = ligado a su Indicador Nivel 2 con codigo = nombre hoja

Genera:
  - JSON con la lista de templates lista para POST /admin/templates/import
  - Plan de SQL para limpiar la BD (lo ejecuta separado)
"""
import json
import re
import openpyxl
from pathlib import Path

XLSX_PATH = Path(__file__).parent / "data" / "PAI_DEFINITIVO_2026.xlsx"
OUT_PATH  = Path(__file__).parent / "data" / "pai_templates.json"

# ── Mapeo de columnas Excel → campos del template ──────────────────────────
# Tuple: (col_idx 1-based, name, label, type, readonly, required, validator_only)
COLS = [
    # constantes del producto (readonly)
    (1,  "periodo_reporte",       "Periodo de Reporte",                    "select",   False, True,  False),
    (2,  "linea_estrategica",     "Línea estratégica",                     "textarea", True,  False, False),
    (3,  "resultado_estrategico", "Resultados Estratégico 2024 - 2028",    "textarea", True,  False, False),
    (4,  "resultado_2026",        "Resultado 2026",                        "textarea", True,  False, False),
    (5,  "codigo_producto",       "Código del Producto",                   "text",     True,  False, False),
    (6,  "producto",              "Producto",                              "textarea", True,  False, False),
    (7,  "objetivo_producto",     "Objetivo del Producto",                 "textarea", True,  False, False),
    (8,  "area_responsable",      "Área Responsable",                      "textarea", True,  False, False),
    (9,  "area_implementadora",   "Área implementadora o corresponsable",  "textarea", True,  False, False),
    # variables de la actividad (editable por la dependencia)
    (10, "eje",                   "Ejes",                                  "text",     False, True,  False),
    (11, "peso_actividad",        "Peso de la actividad en la medición",   "number",   False, True,  False),
    (12, "peso_trimestre",        "Peso para el trimestre",                "number",   False, True,  False),
    (13, "actividad_clave",       "Actividades Clave",                     "textarea", False, True,  False),
    (14, "indicador",             "Indicador",                             "textarea", False, True,  False),
    (15, "meta_anual",            "Meta Anual",                            "textarea", False, True,  False),
    (16, "entregable_total",      "Entregable Total",                      "textarea", False, True,  False),
    (17, "entregable_trimestre",  "Entregable trimestre",                  "textarea", False, True,  False),
    # avances reportados por la dependencia
    (18, "pct_avance_proyectado", "% Avance acumulado Proyectado de la actividad", "number", False, True,  False),
    (19, "pct_avance_alcanzado",  "% Avance acumulado Alcanzado de la actividad",  "number", False, True,  False),
    # Auto-calculados (el frontend y el upload Excel los recalculan siempre)
    (20, "pct_avance_final",      "% Avance final del periodo de la actividad",    "number", True,  False, False),
    (21, "estado_actividad",      "Estado de Cumplimiento de la actividad",        "select", True,  False, False),
    # ponderación a nivel de producto (calculados / readonly)
    (22, "pct_avance_ponderado",  "% Avance trimestral ponderado del producto",     "number", True,  False, False),
    (23, "estado_ponderado",      "Estado de Cumplimiento ponderado del producto",  "select", True,  False, False),
    (24, "pct_avance_anual",      "% Avance anual del producto",                    "number", True,  False, False),
    # comentarios de la dependencia
    (25, "avances_logros",        "Avances y logros",                                "textarea", False, False, False),
    (26, "retrasos_dificultades", "Retrasos o dificultades",                         "textarea", False, False, False),
    (27, "comentarios_extra",     "Comentarios adicionales",                         "textarea", False, False, False),
    # observaciones del validador (OAP) — SOLO las llena el validador al aprobar
    (28, "obs_oap",               "Observaciones y Recomendaciones OAP",                              "textarea", False, True,  True),
    (29, "obs_oap_estado",        "Observaciones OAP respecto al Estado de Cumplimiento ponderado",   "textarea", False, True,  True),
]

PERIODO_OPTS = ["TRIMESTRE 1", "TRIMESTRE 2", "TRIMESTRE 3", "TRIMESTRE 4"]
ESTADO_OPTS  = ["Cumple", "Cumple Parcialmente", "No Cumple", "No Aplica", "Sin Reporte"]


def clean_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_linea_num(s: str | None) -> int | None:
    """Extrae el número de línea estratégica desde 'Línea N. ...'."""
    if not s:
        return None
    m = re.match(r"\s*Línea\s+(\d+)", s)
    return int(m.group(1)) if m else None


def parse_codigo_linea(codigo: str) -> int | None:
    """De 'L3-P1-DPE-2026' extrae el 3."""
    m = re.match(r"^L(\d+)", codigo)
    return int(m.group(1)) if m else None


def build_field(col_idx, name, label, ftype, readonly, required, validator_only,
                default=None, options=None, auto_calculate=None):
    f = {"name": name, "label": label, "type": ftype,
         "readonly": readonly, "required": required,
         "validator_only": validator_only}
    if default is not None:
        f["default"] = default
    if options:
        f["options"] = options
    if auto_calculate:
        f["auto_calculate"] = auto_calculate
    return f


# Reglas de auto-cálculo aplicadas a TODOS los templates del PAI
AUTO_CALC_BY_FIELD = {
    "pct_avance_final":  "ratio_alcanzado_proyectado",
    "estado_actividad":  "estado_cumplimiento_from_pct_final",
}


def hoja_to_template(ws):
    sheet_name = ws.title
    # Recoger valores constantes (toman el valor de la primera fila de datos)
    consts = {}
    for col_idx, name, label, ftype, ro, req, vo in COLS:
        if not ro:
            continue
        # Buscar el primer valor no-vacío en las filas de datos
        for r in range(2, ws.max_row + 1):
            v = clean_text(ws.cell(row=r, column=col_idx).value)
            if v:
                consts[name] = v
                break

    codigo_template = consts.get("codigo_producto") or sheet_name
    producto = consts.get("producto") or sheet_name
    linea_completa = consts.get("linea_estrategica") or ""

    # Campos del template
    fields = []
    for col_idx, name, label, ftype, ro, req, vo in COLS:
        default = consts.get(name) if ro else None
        opts = None
        if name == "periodo_reporte":
            opts = PERIODO_OPTS
        elif name in ("estado_actividad", "estado_ponderado"):
            opts = ESTADO_OPTS
        auto = AUTO_CALC_BY_FIELD.get(name)
        fields.append(build_field(col_idx, name, label, ftype, ro, req, vo,
                                   default=default, options=opts,
                                   auto_calculate=auto))

    # Markdown sencillo: tabla con los campos
    md_lines = [
        f"# {codigo_template}\n",
        f"**Producto:** {producto}\n",
        f"**Línea estratégica:** {linea_completa[:200]}\n\n",
        "| Campo | Etiqueta | Tipo | Bloqueado | Requerido | Solo Validador | Default | Opciones |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for f in fields:
        opts_str = ",".join(f.get("options", [])) if f.get("options") else ""
        default_str = (f.get("default") or "")[:50] if isinstance(f.get("default"), str) else ""
        md_lines.append(
            f"| {f['name']} | {f['label']} | {f['type']} | "
            f"{'Sí' if f['readonly'] else 'No'} | "
            f"{'Sí' if f['required'] else 'No'} | "
            f"{'Sí' if f.get('validator_only') else 'No'} | "
            f"{default_str} | {opts_str} |"
        )
    markdown = "\n".join(md_lines)

    # Nombre del template: corto y legible
    short_name = f"{codigo_template} — {producto[:80]}"

    return {
        "codigo": codigo_template,
        "nombre": short_name,
        "descripcion": (consts.get("objetivo_producto") or producto)[:500],
        "linea_num": parse_codigo_linea(codigo_template),
        "producto_full": producto,
        "linea_full": linea_completa,
        "objetivo": consts.get("objetivo_producto"),
        "codigo_markdown": markdown,
        "configuracion_campos": {"fields": fields},
        "version": 1,
        "activo": True,
    }


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    templates = []
    for sheet_name in wb.sheetnames:
        t = hoja_to_template(wb[sheet_name])
        templates.append(t)
        print(f"✓ {t['codigo']:25s} — Línea {t['linea_num']} — {t['producto_full'][:60]}")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(templates, ensure_ascii=False, indent=2))
    print(f"\n→ Guardado: {OUT_PATH}  ({len(templates)} templates)")

    # Resumen por línea
    from collections import defaultdict
    por_linea = defaultdict(list)
    for t in templates:
        por_linea[t['linea_num']].append(t['codigo'])
    print("\nResumen por Línea Estratégica:")
    for k in sorted(por_linea):
        print(f"  L{k}: {len(por_linea[k])} producto(s) → {por_linea[k]}")


if __name__ == "__main__":
    main()
