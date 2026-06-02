"""
adapt_referencias.py — Convierte los Excels descargados desde el sistema
(`formularios_*_aprobados.xlsx`) a una versión MÍNIMA, lista para
re-cargarse por la UI de "Cargar Excel" de la dependencia.

Qué hace:
  - Mantiene SOLO las columnas que la dependencia debe llenar:
    los campos no-readonly y no-validator_only y no-auto_calculate.
  - Elimina valores readonly y validator/auto (el sistema los inyecta o
    calcula al cargar). Así evita conflictos de "valor inesperado".
  - Si el archivo origen no tiene filas de datos reales (solo headers +
    fila de tipo) lo SALTA (no produce salida) y avisa en consola.
  - Conserva el formato esperado por upload-excel: fila 1 = headers,
    fila 2 = tipo/guía, filas 3+ = datos.

Uso:
    python scripts/pai_2026/adapt_referencias.py
        # entrada:  referencias/formularios_*_aprobados.xlsx
        # salida:   referencias/adaptados/<codigo>_para_cargar.xlsx
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT / "referencias"
OUT_DIR = ROOT / "referencias" / "adaptados"

# ── Cargar definición de campos del template PAI ──────────────────────────────
setup_src = (ROOT / "scripts/pai_2026/setup_pai.py").read_text(encoding="utf-8")
m = re.search(r"COLS\s*=\s*\[(.*?)\n\]\s*\n", setup_src, re.DOTALL)
if not m:
    sys.exit("No pude leer COLS de setup_pai.py")
fields: list[dict] = []
for line in m.group(1).splitlines():
    mt = re.match(
        r'\s*\(\s*\d+\s*,\s*"([^"]+)"\s*,\s*"([^"]+)"\s*,\s*"([^"]+)"\s*,'
        r'\s*(True|False)\s*,\s*(True|False)\s*,\s*(True|False)\s*\)',
        line,
    )
    if not mt:
        continue
    fields.append({
        "name": mt.group(1),
        "label": mt.group(2),
        "type": mt.group(3),
        "readonly": mt.group(4) == "True",
        "required": mt.group(5) == "True",
        "validator_only": mt.group(6) == "True",
        "auto_calculate": mt.group(1) in {"pct_avance_final", "estado_actividad"},
    })

# Campos llenables (los que la dependencia DEBE editar)
FILLABLE = [f for f in fields if not f["readonly"]
            and not f["validator_only"] and not f["auto_calculate"]]
FILLABLE_LABELS = [f["label"] for f in FILLABLE]
FILLABLE_BY_LABEL = {f["label"]: f for f in FILLABLE}


def _type_hint(f: dict) -> str:
    t = f["type"]
    req = "* " if f["required"] else ""
    if t == "number":   return req + "Número"
    if t == "date":     return req + "Fecha (AAAA-MM-DD)"
    if t == "textarea": return req + "Texto largo"
    if t == "select":
        # Para periodo_reporte sabemos las opciones
        if f["name"] == "periodo_reporte":
            return req + "Opciones: TRIMESTRE 1 / TRIMESTRE 2 / TRIMESTRE 3 / TRIMESTRE 4"
        return req + "Seleccionar"
    return req + "Texto"


def _is_type_hint_row(row: tuple) -> bool:
    keywords = {"número", "texto", "fecha", "opciones", "solo lectura",
                "texto largo", "seleccionar"}
    vals = [str(v).strip().lower() if v is not None else "" for v in row]
    return any(any(kw in v for kw in keywords) for v in vals if v)


def adapt(src_path: Path) -> Path | None:
    wb_in = openpyxl.load_workbook(src_path, data_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]
    all_rows = list(ws_in.iter_rows(values_only=True))
    if len(all_rows) < 2:
        print(f"  ⚠️  {src_path.name}: tiene {len(all_rows)} fila(s) → vacío, se omite")
        return None

    headers_in = [str(h).strip() if h else "" for h in all_rows[0]]
    # Detectar y saltar fila tipo
    data_start = 1
    if _is_type_hint_row(all_rows[1]):
        data_start = 2

    # Índice por header → columna del Excel origen
    src_col_by_label = {h: i for i, h in enumerate(headers_in) if h}

    data_rows_in = all_rows[data_start:]
    real_rows: list[tuple] = []
    for row in data_rows_in:
        non_empty = [c for c in row if c is not None and str(c).strip() not in ("", "None")]
        if not non_empty:
            continue
        first = str(row[0]).strip().lower() if row[0] else ""
        if first.startswith("leyenda"):
            continue
        real_rows.append(row)

    if not real_rows:
        print(f"  ⚠️  {src_path.name}: no tiene filas de datos reales → se omite")
        return None

    # Construir Excel de salida con SOLO los campos llenables
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ws_in.title[:31]

    # Estilos
    header_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    type_fill   = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    type_font   = Font(color="1A3C5E", italic=True, size=9)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Fila 1: headers
    ws.append(FILLABLE_LABELS)
    # Fila 2: tipo/guía
    ws.append([_type_hint(f) for f in FILLABLE])

    # Filas 3+: datos extraídos
    for row in real_rows:
        out_row = []
        for f in FILLABLE:
            src_idx = src_col_by_label.get(f["label"])
            v = row[src_idx] if (src_idx is not None and src_idx < len(row)) else None
            out_row.append(v)
        ws.append(out_row)

    # Estilo a fila 1 y 2
    for col_idx in range(1, len(FILLABLE) + 1):
        c1 = ws.cell(row=1, column=col_idx)
        c1.fill, c1.font, c1.alignment = header_fill, header_font, center
        c2 = ws.cell(row=2, column=col_idx)
        c2.fill, c2.font, c2.alignment = type_fill, type_font, left
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 28

    ws.freeze_panes = "A3"

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    # Nombre: <codigo>_para_cargar.xlsx
    code = src_path.stem.replace("formularios_", "").replace("_aprobados", "")
    out_path = OUT_DIR / f"{code}_para_cargar.xlsx"
    wb.save(out_path)
    print(f"  ✅ {src_path.name}  →  {out_path.relative_to(ROOT)}  ({len(real_rows)} filas)")
    return out_path


def main() -> int:
    files = sorted(SRC_DIR.glob("formularios_*_aprobados.xlsx"))
    if not files:
        print(f"No hay archivos formularios_*_aprobados.xlsx en {SRC_DIR}")
        return 1
    print(f"Adaptando {len(files)} archivo(s) → {OUT_DIR}/\n")
    ok = empty = 0
    for f in files:
        if adapt(f) is None:
            empty += 1
        else:
            ok += 1
    print(f"\nResumen: {ok} adaptados | {empty} omitidos (vacíos)")
    print(f"Carpeta de salida: {OUT_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
