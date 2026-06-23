"""
scripts/migrate_datos_formularios_2.py

Migración puntual:
  1. ELIMINA todos los formularios existentes (datos de prueba) + adjuntos MinIO.
  2. Carga los Excel de DATOS_FORMULARIOS_2 usando los TEMPLATES ACTUALES
     (mapeo por label de columna → nombre de campo del template vigente).
  3. Ignora los Excel que no tienen filas de datos.
  4. Crea los formularios en estado APPROVED (eran exportes "_aprobados"),
     asociados a la dependencia indicada en la hoja "Referencia".

Se ejecuta dentro del contenedor backend (tiene acceso a la BD y a los modelos).
Uso:
    python scripts/migrate_datos_formularios_2.py /ruta/a/DATOS_FORMULARIOS_2
"""
from __future__ import annotations

import asyncio
import glob
import os
import sys
from datetime import date, datetime, timezone

import openpyxl
from sqlalchemy import delete as sa_delete, select

# Importar app
sys.path.insert(0, "/app")

from app.database import AsyncSessionLocal  # type: ignore
from app.models.dependency import Dependency  # type: ignore
from app.models.file import Archivo  # type: ignore
from app.models.form import Form, FormStatus  # type: ignore
from app.models.template import Template  # type: ignore
from app.models.user import User, UserRole  # type: ignore
from app.routers.forms_excel import (  # type: ignore
    _field_label,
    _field_name,
    _field_type,
    _get_fields,
    _is_readonly,
)
from app.services.auto_calc import recalc_auto_fields  # type: ignore
from app.services.minio_service import get_minio_service  # type: ignore

TYPE_HINT_KEYWORDS = {"número", "texto", "fecha", "opciones", "solo lectura", "texto largo", "seleccionar"}


def _str_val(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "none" else s


def _norm(s: str) -> str:
    return (s or "").strip().lower()


async def wipe_existing(db) -> dict:
    """Borra todos los forms + archivos (datos de prueba)."""
    minio = get_minio_service()
    archivos = (await db.execute(select(Archivo))).scalars().all()
    minio_ok = minio_fail = 0
    for a in archivos:
        try:
            minio.delete_file(a)
            minio_ok += 1
        except Exception as exc:  # noqa: BLE001
            print(f"  [warn] no se pudo borrar MinIO {a.id}: {exc}")
            minio_fail += 1
    del_arch = (await db.execute(sa_delete(Archivo))).rowcount or 0
    del_forms = (await db.execute(sa_delete(Form))).rowcount or 0
    await db.commit()
    return {"forms": del_forms, "archivos": del_arch, "minio_ok": minio_ok, "minio_fail": minio_fail}


async def load_excel(db, path: str, dep_index: dict, default_dep, creator: User, validator: User | None) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    main = wb[wb.sheetnames[0]]
    codigo = wb.sheetnames[0].strip()

    rows = list(main.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"codigo": codigo, "skipped": True, "reason": "sin filas", "created": 0}

    headers = [_str_val(h) for h in rows[0]]

    # Detectar fila de tipo/guía
    data_start = 1
    row2 = [_norm(_str_val(v)) for v in rows[1]] if len(rows) >= 2 else []
    if any(any(kw in v for kw in TYPE_HINT_KEYWORDS) for v in row2 if v):
        data_start = 2
    data_rows = rows[data_start:]

    # ¿Hay filas con datos?
    real_rows = [
        r for r in data_rows
        if any(c is not None and str(c).strip() not in ("", "None") for c in r)
        and not _norm(_str_val(r[0] if r else None)).startswith("leyenda")
    ]
    if not real_rows:
        return {"codigo": codigo, "skipped": True, "reason": "sin datos", "created": 0}

    # Buscar template ACTUAL por codigo
    tpl = (await db.execute(select(Template).where(Template.codigo == codigo))).scalar_one_or_none()
    if tpl is None:
        return {"codigo": codigo, "skipped": True, "reason": "template no existe", "created": 0}

    all_fields = _get_fields(tpl)
    field_by_label = {_field_label(f): f for f in all_fields}
    field_by_name = {_field_name(f): f for f in all_fields}

    # Mapa columna → field (ignora validator_only / auto_calculate / extras)
    col_map: dict[int, dict] = {}
    for idx, h in enumerate(headers):
        if not h:
            continue
        f = field_by_label.get(h) or field_by_name.get(h)
        if f is None:
            continue
        if _field_name(f) == "informe_cualitativo":
            continue
        if f.get("validator_only") or f.get("auto_calculate"):
            continue
        col_map[idx] = f

    # Dependencia (de la hoja Referencia, por fila). Fallback: default_dep.
    dep_by_row: dict[int, object] = {}
    if "Referencia" in wb.sheetnames:
        for r in list(wb["Referencia"].iter_rows(values_only=True))[1:]:
            try:
                fila = int(str(r[0]).strip())
            except (TypeError, ValueError):
                continue
            dep_name = _str_val(r[3]) if len(r) > 3 else ""
            dep = dep_index.get(_norm(dep_name)) if dep_name else None
            if dep:
                dep_by_row[fila] = dep

    created = 0
    for pos, row in enumerate(real_rows, start=1):
        datos: dict = {}
        for idx, f in col_map.items():
            raw = row[idx] if idx < len(row) else None
            val = _str_val(raw)
            fname = _field_name(f)
            ftype = _field_type(f)
            readonly = _is_readonly(f)
            if ftype == "computed":
                continue
            if not val and readonly and f.get("default") is not None:
                datos[fname] = f["default"]
            elif ftype == "number" and val:
                try:
                    datos[fname] = float(str(raw).replace(",", "."))
                except (TypeError, ValueError):
                    datos[fname] = None
            elif ftype == "date" and raw is not None:
                if isinstance(raw, (date, datetime)):
                    datos[fname] = raw.isoformat() if isinstance(raw, date) and not isinstance(raw, datetime) else raw.date().isoformat()
                else:
                    datos[fname] = val or None
            else:
                datos[fname] = val or None

        # defaults readonly faltantes
        for f in all_fields:
            fn = _field_name(f)
            if _is_readonly(f) and f.get("default") is not None and not f.get("auto_calculate"):
                datos.setdefault(fn, f["default"])

        recalc_auto_fields(datos, all_fields)

        dep = dep_by_row.get(pos) or default_dep
        form = Form(
            plantilla_id=tpl.id,
            usuario_id=creator.id,
            dependency_id=dep.id,
            datos_dinamicos=datos,
            informe_cualitativo=None,
            fecha_usuario=date.today(),
            estado=FormStatus.approved,
            cargado_via_excel=True,
            validado_por_id=validator.id if validator else None,
            fecha_validacion=datetime.now(timezone.utc),
        )
        db.add(form)
        created += 1

    await db.commit()
    return {"codigo": codigo, "skipped": False, "created": created, "dep": getattr(default_dep, "codigo", "?")}


async def main(folder: str):
    async with AsyncSessionLocal() as db:
        # Resolver dependencias por nombre
        deps = (await db.execute(select(Dependency))).scalars().all()
        dep_index = {_norm(d.nombre): d for d in deps}
        # Default dep: OAP001 (Oficina Asesora de Planeacion)
        default_dep = next((d for d in deps if d.codigo == "OAP001"), deps[0] if deps else None)
        if default_dep is None:
            print("ERROR: no hay dependencias en la BD.")
            return

        # Usuario creador: admin. Validador: rol validator si existe.
        admin = (await db.execute(select(User).where(User.role == UserRole.admin))).scalars().first()
        validator = (await db.execute(select(User).where(User.role == UserRole.validator))).scalars().first()
        creator = admin or validator
        if creator is None:
            print("ERROR: no hay usuario admin/validator.")
            return

        print(f"Creador: {creator.username} | Dep por defecto: {default_dep.codigo} | Validador: {validator.username if validator else '—'}")

        # 1) Wipe
        print("\n== Eliminando registros de prueba existentes ==")
        wiped = await wipe_existing(db)
        print(f"  Borrados: {wiped['forms']} forms, {wiped['archivos']} archivos "
              f"(MinIO ok={wiped['minio_ok']}, fail={wiped['minio_fail']})")

        # 2) Cargar Excels
        print("\n== Cargando DATOS_FORMULARIOS_2 ==")
        total = 0
        files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
        for path in files:
            res = await load_excel(db, path, dep_index, default_dep, creator, validator)
            if res["skipped"]:
                print(f"  [skip] {res['codigo']:18s} — {res['reason']}")
            else:
                total += res["created"]
                print(f"  [ok]   {res['codigo']:18s} — {res['created']} registros (dep {res.get('dep')})")

        print(f"\nTOTAL creados: {total} formularios APROBADOS")


if __name__ == "__main__":
    folder = sys.argv[1] if len(sys.argv) > 1 else "/data/DATOS_FORMULARIOS_2"
    asyncio.run(main(folder))
