"""
app/routers/stats.py — Endpoints de estadísticas públicas (sin autenticación).
"""
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.dependency import Dependency
from app.models.fact_stats import FactStats
from app.models.form import Form, FormStatus
from app.models.indicator import Indicator
from app.models.template import Template
from app.models.user import User
from app.schemas.stats import (
    DetailedFormResponse,
    GlobalStatsItem,
    GlobalStatsResponse,
    IndicatorResponse,
    TemplateStatsItem,
    TemplateStatsResponse,
)

router = APIRouter(prefix="/stats", tags=["Estadísticas Públicas"])


def _apply_date_filter(query, model, start_date: Optional[date], end_date: Optional[date]):
    """Aplica filtros de fecha a la columna fecha_aprobacion de FactStats."""
    if start_date:
        query = query.where(model.fecha_aprobacion >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.where(model.fecha_aprobacion <= datetime.combine(end_date, datetime.max.time()))
    return query


@router.get("/indicators", response_model=list[IndicatorResponse])
async def list_indicators(
    db: AsyncSession = Depends(get_db),
) -> list[IndicatorResponse]:
    """Lista todos los indicadores de Nivel 1 activos."""
    result = await db.execute(
        select(Indicator).where(Indicator.activo == True).order_by(Indicator.id)
    )
    indicators = result.scalars().all()
    return [IndicatorResponse.model_validate(i) for i in indicators]


@router.get("/global", response_model=GlobalStatsResponse)
async def get_global_stats(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
) -> GlobalStatsResponse:
    """
    Nivel 1: completitud promedio por indicador en un rango de fechas.
    Endpoint público sin autenticación.
    """
    query = (
        select(
            Indicator.id,
            Indicator.nombre,
            func.avg(FactStats.completitud).label("completitud_promedio"),
            func.count(FactStats.id).label("total_formularios"),
        )
        .join(FactStats, FactStats.indicador_id == Indicator.id)
        .where(Indicator.activo == True)
        .group_by(Indicator.id, Indicator.nombre)
        .order_by(Indicator.id)
    )
    query = _apply_date_filter(query, FactStats, start_date, end_date)
    result = await db.execute(query)
    rows = result.all()

    items = [
        GlobalStatsItem(
            indicador_id=row.id,
            nombre=row.nombre,
            completitud_promedio=round(float(row.completitud_promedio or 0), 2),
            total_formularios=row.total_formularios,
            start_date=start_date,
            end_date=end_date,
        )
        for row in rows
    ]
    return GlobalStatsResponse(items=items, start_date=start_date, end_date=end_date)


@router.get("/by-template", response_model=TemplateStatsResponse)
async def get_stats_by_template(
    indicador_id: int = Query(..., description="ID del indicador de Nivel 1"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
) -> TemplateStatsResponse:
    """
    Nivel 2: completitud por template para un indicador específico.
    """
    # Obtener nombre del indicador
    ind_result = await db.execute(
        select(Indicator).where(Indicator.id == indicador_id)
    )
    indicator = ind_result.scalar_one_or_none()
    indicador_nombre = indicator.nombre if indicator else f"Indicador {indicador_id}"

    query = (
        select(
            Template.id,
            Template.nombre,
            func.avg(FactStats.completitud).label("completitud"),
            func.count(FactStats.id).label("total_formularios"),
        )
        .join(FactStats, FactStats.template_id == Template.id)
        .where(FactStats.indicador_id == indicador_id)
        .group_by(Template.id, Template.nombre)
        .order_by(Template.nombre)
    )
    query = _apply_date_filter(query, FactStats, start_date, end_date)
    result = await db.execute(query)
    rows = result.all()

    items = [
        TemplateStatsItem(
            template_id=row.id,
            nombre=row.nombre,
            completitud=round(float(row.completitud or 0), 2),
            total_formularios=row.total_formularios,
        )
        for row in rows
    ]
    return TemplateStatsResponse(
        indicador_id=indicador_id,
        indicador_nombre=indicador_nombre,
        items=items,
        start_date=start_date,
        end_date=end_date,
    )


_PERIODO_TRIM_MAP = {
    "trim1": "TRIMESTRE 1",
    "trim2": "TRIMESTRE 2",
    "trim3": "TRIMESTRE 3",
    "trim4": "TRIMESTRE 4",
}


@router.get("/detail", response_model=DetailedFormResponse)
async def get_stats_detail(
    template_id: str = Query(...),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    periodo: Optional[str] = Query(None, description="anual | trim1 | trim2 | trim3 | trim4"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
) -> DetailedFormResponse:
    """
    Nivel 3: lista de formularios validados para un template dado.
    """
    import uuid as _uuid
    from sqlalchemy import cast, String
    from sqlalchemy import or_

    try:
        tmpl_uuid = _uuid.UUID(template_id)
    except ValueError:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="template_id inválido")

    query = (
        select(Form, Dependency.nombre.label("dep_nombre"), User.nombre_completo.label("user_nombre"))
        .join(Dependency, Form.dependency_id == Dependency.id)
        .join(User, Form.usuario_id == User.id)
        .where(Form.plantilla_id == tmpl_uuid, Form.estado == FormStatus.approved)
    )

    if start_date:
        query = query.where(Form.fecha_validacion >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.where(Form.fecha_validacion <= datetime.combine(end_date, datetime.max.time()))

    trim_value = _PERIODO_TRIM_MAP.get((periodo or "").lower())
    if trim_value:
        query = query.where(Form.datos_dinamicos["periodo_reporte"].astext == trim_value)

    if search:
        query = query.where(
            or_(
                Dependency.nombre.ilike(f"%{search}%"),
                User.nombre_completo.ilike(f"%{search}%"),
                cast(Form.datos_dinamicos, String).ilike(f"%{search}%"),
            )
        )

    count_q = select(func.count()).select_from(query.subquery())
    total = await db.scalar(count_q) or 0

    offset = (page - 1) * size
    result = await db.execute(
        query.order_by(Form.fecha_validacion.desc()).offset(offset).limit(size)
    )
    rows = result.all()

    from app.schemas.stats import DetailedFormItem

    items = [
        DetailedFormItem(
            id=row.Form.id,
            fecha_referencia=row.Form.fecha_usuario,
            fecha_carga=row.Form.fecha_carga,
            dependencia=row.dep_nombre,
            usuario=row.user_nombre,
            informe_cualitativo=row.Form.informe_cualitativo,
            datos_dinamicos=row.Form.datos_dinamicos,
            archivos_count=len(row.Form.archivos),
        )
        for row in rows
    ]

    return DetailedFormResponse(total=total, page=page, size=size, items=items)


@router.get("/export")
async def export_stats_excel(
    template_id: str = Query(...),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    periodo: Optional[str] = Query(None, description="anual | trim1..4"),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Exporta los formularios validados del Nivel 3 como archivo Excel (.xlsx).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import uuid as _uuid

    try:
        tmpl_uuid = _uuid.UUID(template_id)
    except ValueError:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="template_id inválido")

    # Reusar la lógica de detalle sin paginación
    detail = await get_stats_detail(
        template_id=template_id,
        start_date=start_date,
        end_date=end_date,
        periodo=periodo,
        page=1,
        size=10000,
        search=search,
        db=db,
    )

    # Obtener nombre del template
    tmpl_result = await db.execute(select(Template).where(Template.id == tmpl_uuid))
    template = tmpl_result.scalar_one_or_none()
    tmpl_nombre = template.nombre if template else template_id

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estadísticas"

    # Encabezado
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    headers = [
        "ID", "Fecha Referencia", "Fecha Carga", "Dependencia",
        "Usuario", "Informe Cualitativo", "Archivos Adjuntos"
    ]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Filas
    for row_num, item in enumerate(detail.items, start=2):
        ws.cell(row=row_num, column=1, value=str(item.id))
        ws.cell(row=row_num, column=2, value=str(item.fecha_referencia) if item.fecha_referencia else "")
        ws.cell(row=row_num, column=3, value=item.fecha_carga.strftime("%Y-%m-%d %H:%M") if item.fecha_carga else "")
        ws.cell(row=row_num, column=4, value=item.dependencia)
        ws.cell(row=row_num, column=5, value=item.usuario)
        ws.cell(row=row_num, column=6, value=item.informe_cualitativo or "")
        ws.cell(row=row_num, column=7, value=item.archivos_count)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 60)

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"estadisticas_{tmpl_nombre.replace(' ', '_')}"
    if start_date:
        filename += f"_{start_date}"
    if end_date:
        filename += f"_a_{end_date}"
    filename += ".xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
