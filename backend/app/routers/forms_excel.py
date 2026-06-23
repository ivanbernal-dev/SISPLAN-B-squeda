"""
app/routers/forms_excel.py — Endpoints para carga y descarga de formularios via Excel.

La clave interna de cada campo en configuracion_campos es "type" (no "tipo").
La carga valida campos requeridos, tipos numéricos y opciones de select.
"""
import io
import logging
import uuid
from datetime import date
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_admin_user, get_any_authenticated, get_dependency_user
from app.models.file import Archivo
from app.models.form import Form, FormStatus
from app.models.template import Template
from app.models.user import User
from app.services.minio_service import get_minio_service

router = APIRouter(tags=["Formularios Excel"])
logger = logging.getLogger(__name__)


def _get_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl no está instalado. Ejecute: pip install openpyxl",
        ) from e


def _get_fields(template: Template) -> list[dict]:
    """
    Retorna la lista de campos del template.
    Soporta tanto {"fields": [...]} como {"campos": [...]}.
    Excluye campos de tipo 'computed' (se calculan automáticamente).
    """
    cfg = template.configuracion_campos or {}
    fields = cfg.get("fields") or cfg.get("campos") or []
    return [f for f in fields if f.get("type", f.get("tipo", "text")) not in ("computed", "archivos", "file")]


def _field_type(f: dict) -> str:
    """Obtiene el tipo normalizado del campo (acepta 'type' o 'tipo')."""
    return str(f.get("type", f.get("tipo", "text"))).lower()


def _field_options(f: dict) -> list:
    """Obtiene las opciones de un campo select (acepta 'options' u 'opciones')."""
    return f.get("options") or f.get("opciones") or []


def _field_label(f: dict) -> str:
    return str(f.get("label", f.get("name", ""))).strip()


def _field_name(f: dict) -> str:
    return str(f.get("name", "")).strip()


def _is_readonly(f: dict) -> bool:
    return bool(f.get("readonly", f.get("bloqueado", False)))


def _is_required(f: dict) -> bool:
    return bool(f.get("required", f.get("requerido", False)))


# ── GET /templates/{template_id}/excel-example ───────────────────────────────

@router.get("/templates/{template_id}/excel-example")
async def download_excel_example(
    template_id: uuid.UUID,
    current_user: User = Depends(get_any_authenticated),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Descarga un Excel de ejemplo para el template.
    Incluye todos los campos editables + informe_cualitativo.
    Fila 1: Encabezados con color.
    Fila 2: Tipo de dato esperado (guía).
    Fila 3: Ejemplo de valores.
    """
    openpyxl = _get_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    result = await db.execute(
        select(Template).where(Template.id == template_id, Template.activo == True)
    )
    template = result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template no encontrado")

    fields = _get_fields(template)

    # Columnas a exportar: solo los campos que la dependencia debe rellenar.
    # NO se exportan los validator_only (los llena el validador) ni los
    # auto_calculate (los recalcula el sistema al guardar / cargar).
    export_cols: list[dict] = [
        f for f in fields
        if not f.get("validator_only") and not f.get("auto_calculate")
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (template.codigo or template.nombre)[:31]

    # ── Estilos ──────────────────────────────────────────────────────────────
    header_fill    = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    required_fill  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # verde claro
    optional_fill  = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")  # amarillo claro
    readonly_fill  = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")  # gris
    type_row_fill  = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # azul muy claro

    header_font  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    type_font    = Font(color="1A3C5E", italic=True, size=9, name="Calibri")
    normal_font  = Font(size=10, name="Calibri")
    req_font     = Font(color="1B5E20", size=10, name="Calibri")
    opt_font     = Font(color="5D4037", size=10, name="Calibri")
    ro_font      = Font(color="757575", size=10, name="Calibri")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Fila 1: Encabezados ───────────────────────────────────────────────────
    headers = [_field_label(f) for f in export_cols]
    ws.append(headers)

    # ── Fila 2: Tipo de dato / guía ────────────────────────────────────────
    type_hints = []
    for f in export_cols:
        ft = _field_type(f)
        ro = _is_readonly(f)
        req = _is_required(f)

        if ro:
            hint = "Solo lectura"
        elif ft == "number":
            hint = "Número"
        elif ft == "date":
            hint = "Fecha (AAAA-MM-DD)"
        elif ft == "select":
            opts = _field_options(f)
            opts_str = " / ".join(str(o.get("value", o) if isinstance(o, dict) else o) for o in opts[:5])
            hint = f"Opciones: {opts_str}" if opts_str else "Seleccionar"
        elif ft == "textarea":
            hint = "Texto largo"
        else:
            hint = "Texto"

        if not ro:
            hint = ("* " if req else "") + hint
        type_hints.append(hint)
    ws.append(type_hints)

    # ── Fila 3: Ejemplo de valores ────────────────────────────────────────
    sample_row = []
    for f in export_cols:
        ft = _field_type(f)
        default = f.get("default")
        if _is_readonly(f):
            sample_row.append(str(default) if default is not None else "")
        elif ft == "date":
            sample_row.append(date.today().isoformat())
        elif ft == "number":
            sample_row.append(0)
        elif ft == "select":
            opts = _field_options(f)
            if opts:
                first = opts[0]
                sample_row.append(first.get("value", first) if isinstance(first, dict) else first)
            else:
                sample_row.append("")
        elif _field_name(f) == "informe_cualitativo":
            sample_row.append("Descripción detallada de los hallazgos y procesos de búsqueda.")
        else:
            sample_row.append(str(default) if default is not None else "Ejemplo de valor")
    ws.append(sample_row)

    # ── Aplicar estilos columna por columna ───────────────────────────────
    for col_idx, f in enumerate(export_cols, start=1):
        col_letter = get_column_letter(col_idx)
        ro = _is_readonly(f)
        req = _is_required(f)

        # Ancho de columna
        ws.column_dimensions[col_letter].width = 28

        # Fila 1: header
        h_cell = ws.cell(row=1, column=col_idx)
        h_cell.font = header_font
        h_cell.fill = header_fill
        h_cell.alignment = center_align
        h_cell.border = thin_border

        # Fila 2: tipo
        t_cell = ws.cell(row=2, column=col_idx)
        t_cell.font = type_font
        t_cell.fill = type_row_fill
        t_cell.alignment = left_align
        t_cell.border = thin_border

        # Fila 3: ejemplo
        s_cell = ws.cell(row=3, column=col_idx)
        if ro:
            s_cell.fill = readonly_fill
            s_cell.font = ro_font
        elif req:
            s_cell.fill = required_fill
            s_cell.font = req_font
        else:
            s_cell.fill = optional_fill
            s_cell.font = opt_font
        s_cell.alignment = left_align
        s_cell.border = thin_border

    # ── Fila de leyenda ────────────────────────────────────────────────────
    ws.append([""])
    ws.append(["LEYENDA:",
               "Verde = campo requerido (*)",
               "Amarillo = campo opcional",
               "Gris = solo lectura (no editar)",
               "* El símbolo * en la fila de tipo indica campo obligatorio"])
    legend_row = ws.max_row
    legend_cell = ws.cell(row=legend_row, column=1)
    legend_cell.font = Font(bold=True, size=9, color="444444", name="Calibri")
    for col in range(2, 6):
        c = ws.cell(row=legend_row, column=col)
        c.font = Font(size=9, color="666666", name="Calibri")

    # ── Fijar encabezados y zoom ──────────────────────────────────────────
    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 24
    ws.sheet_view.zoomScale = 110

    # ── Hoja de referencia ────────────────────────────────────────────────
    ws_ref = wb.create_sheet("Referencia")
    ws_ref.append(["Campo", "Nombre interno", "Tipo", "Requerido", "Solo lectura", "Opciones válidas"])
    ref_header_font = Font(bold=True, color="FFFFFF", size=10)
    ref_header_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    for cell in ws_ref[1]:
        cell.font = ref_header_font
        cell.fill = ref_header_fill
        cell.alignment = center_align

    for f in export_cols:
        opts = _field_options(f)
        opts_str = ", ".join(str(o.get("value", o) if isinstance(o, dict) else o) for o in opts)
        ws_ref.append([
            _field_label(f),
            _field_name(f),
            _field_type(f),
            "Sí" if _is_required(f) else "No",
            "Sí" if _is_readonly(f) else "No",
            opts_str,
        ])
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws_ref.column_dimensions[col].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"plantilla_{template.codigo or str(template_id)}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── POST /forms/upload-excel/{template_id} ───────────────────────────────────

@router.post(
    "/forms/upload-excel/{template_id}",
    status_code=status.HTTP_201_CREATED,
)
async def upload_excel_forms(
    template_id: uuid.UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(get_dependency_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Carga un Excel con múltiples filas y crea formularios en estado draft.

    Validaciones por fila:
    - Campos requeridos no vacíos
    - Campos numéricos parseable como float
    - Campos select con valor dentro de las opciones permitidas
    - informe_cualitativo obligatorio

    Si hay errores retorna 422 con detalle por fila.

    Cada intento queda registrado en logs/backend/uploads/upload_<ts>_<id>.log
    con el detalle de qué fila falló y por qué (incluso si la respuesta HTTP
    es 400/422), para auditar uploads sin tener que reproducir el caso.
    """
    openpyxl = _get_openpyxl()
    # ── Logger por intento de upload ──────────────────────────────────────
    import logging.handlers as _logh
    from pathlib import Path as _Path
    from datetime import datetime as _dt
    import uuid as _uuid
    _uplog_dir = _Path("/app/logs/uploads")
    try:
        _uplog_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        _uplog_dir = _Path("/tmp/ubpd_uploads")
        _uplog_dir.mkdir(parents=True, exist_ok=True)
    _upid = _uuid.uuid4().hex[:8]
    _upfile = _uplog_dir / f"upload_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}_{_upid}.log"
    _uplog = logging.getLogger(f"ubpd.upload.{_upid}")
    _uplog.setLevel(logging.INFO)
    _uph = logging.FileHandler(_upfile, encoding="utf-8")
    _uph.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s",
                                        "%Y-%m-%d %H:%M:%S"))
    _uplog.addHandler(_uph)
    _uplog.info("═══ UPLOAD-EXCEL start | id=%s | user=%s | template=%s | file=%s ═══",
                _upid, current_user.username or current_user.email,
                template_id, file.filename)

    # Validar tipo MIME
    allowed_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
        "application/zip",  # algunos sistemas envían .xlsx como zip
    }
    if file.content_type not in allowed_types and not (file.filename or "").endswith((".xlsx", ".xls")):
        _uplog.error("400 — tipo MIME no aceptado: %s (filename=%s)",
                     file.content_type, file.filename)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se aceptan archivos Excel (.xlsx / .xls). Se recibió: {file.content_type} ({file.filename})",
        )

    # Cargar template
    t_result = await db.execute(
        select(Template).where(Template.id == template_id, Template.activo == True)
    )
    template = t_result.scalar_one_or_none()
    if template is None:
        _uplog.error("404 — template %s no encontrado o inactivo", template_id)
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template no encontrado")

    all_fields = _get_fields(template)
    _uplog.info("Template OK: %s (%s), %d campos definidos",
                template.codigo or template.nombre, template_id, len(all_fields))

    # Campos disponibles por label y por name
    field_by_label: dict[str, dict] = {_field_label(f): f for f in all_fields}
    field_by_name:  dict[str, dict] = {_field_name(f): f  for f in all_fields}

    # informe_cualitativo puede no estar en los campos del template
    INFORME_LABEL = "Informe cualitativo de la búsqueda"

    # Leer Excel
    content = await file.read()
    _uplog.info("Archivo recibido: %d bytes", len(content))
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        _uplog.exception("400 — openpyxl no pudo leer el .xlsx")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se pudo leer el archivo Excel: {e}",
        )

    all_rows = list(ws.iter_rows(values_only=True))
    _uplog.info("Hoja '%s': %d filas leídas, %d columnas",
                ws.title, len(all_rows), len(all_rows[0]) if all_rows else 0)

    # Debe tener al menos fila 1 (headers) + fila 3 (datos) — saltamos fila 2 (tipo/guía)
    if len(all_rows) < 2:
        _uplog.error("400 — archivo con %d fila(s) (mínimo 2)", len(all_rows))
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El archivo solo tiene {len(all_rows)} fila(s). Necesita encabezados + al menos una fila de datos.",
        )

    header_row = all_rows[0]
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    # Detectar si fila 2 es la fila de tipo/guía (no datos reales)
    # La detectamos si la segunda fila contiene "Número", "Texto", "Fecha", "Solo lectura" etc.
    data_start = 1
    if len(all_rows) >= 2:
        row2_values = [str(v).strip().lower() if v is not None else "" for v in all_rows[1]]
        type_hint_keywords = {"número", "texto", "fecha", "opciones", "solo lectura", "texto largo", "seleccionar"}
        if any(any(kw in v for kw in type_hint_keywords) for v in row2_values if v):
            data_start = 2  # saltar fila de tipo/guía

    data_rows = all_rows[data_start:]

    # Mapear columnas del Excel → campo del template
    # col_map: índice_columna → field_dict (o None para informe_cualitativo)
    col_map: dict[int, dict | str] = {}
    informe_col: int | None = None

    for col_idx, header in enumerate(headers):
        h = header.strip()
        if not h:
            continue
        # Detectar informe_cualitativo por label exacto o por name (legado)
        mapped_f = field_by_label.get(h) or field_by_name.get(h)
        is_informe = (
            h == INFORME_LABEL
            or h.lower() == "informe_cualitativo"
            or (mapped_f is not None and _field_name(mapped_f) == "informe_cualitativo")
        )
        if is_informe:
            informe_col = col_idx
            continue
        # Las columnas que coincidan con campos validator_only o auto_calculate
        # se IGNORAN: el validador llenará los OAP al aprobar y los calculados
        # los recalcula el sistema. No se procesan, no se validan.
        if mapped_f is not None and (mapped_f.get("validator_only") or mapped_f.get("auto_calculate")):
            continue
        # Columnas extra del Excel que no corresponden a ningún campo del template
        # se IGNORAN silenciosamente (no error).
        if h in field_by_label:
            col_map[col_idx] = field_by_label[h]
        elif h in field_by_name:
            col_map[col_idx] = field_by_name[h]

    # ¿El template tiene informe_cualitativo definido como campo?
    template_has_informe_field = any(
        _field_name(f) == "informe_cualitativo" for f in all_fields
    )

    # ── Validar todas las filas antes de crear nada ───────────────────────
    validation_errors: list[dict] = []

    def _cell_val(row: tuple, idx: int) -> Any:
        if idx >= len(row):
            return None
        v = row[idx]
        return v

    def _str_val(v: Any) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        # Eliminar "None" literal que a veces genera openpyxl
        return "" if s.lower() == "none" else s

    for row_idx, row in enumerate(data_rows, start=1):
        # Saltar filas de leyenda / completamente vacías
        non_empty = [c for c in row if c is not None and str(c).strip() not in ("", "None")]
        if not non_empty:
            continue

        # Detectar filas de leyenda por palabra clave
        first_val = _str_val(row[0] if row else None).lower()
        if first_val in {"leyenda:", "nota:", "nota"} or first_val.startswith("leyenda"):
            continue

        row_errors: list[str] = []

        # informe_cualitativo es opcional siempre (eliminado del flujo del PAI)
        informe_val = ""
        if informe_col is not None:
            informe_val = _str_val(_cell_val(row, informe_col))

        # Validar campos mapeados
        for col_idx, field_def in col_map.items():
            if not isinstance(field_def, dict):
                continue
            raw = _cell_val(row, col_idx)
            val = _str_val(raw)
            fname = _field_name(field_def)
            flabel = _field_label(field_def)
            ftype = _field_type(field_def)
            readonly = _is_readonly(field_def)
            required = _is_required(field_def)

            # No validar campos readonly, ni validator_only (los llena el validador),
            # ni auto_calculate (los recalcula el sistema)
            if readonly or field_def.get("validator_only") or field_def.get("auto_calculate"):
                continue

            # Requerido
            if required and not val:
                row_errors.append(f"El campo '{flabel}' es obligatorio")
                continue

            if not val:
                continue  # campo vacío y no requerido — OK

            # Tipo numérico
            if ftype == "number":
                try:
                    float(str(raw).replace(",", "."))
                except (TypeError, ValueError):
                    row_errors.append(f"El campo '{flabel}' debe ser un número (se encontró: '{val}')")

            # Tipo fecha
            elif ftype == "date":
                from datetime import datetime as _dt
                parsed_ok = False
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                    try:
                        _dt.strptime(val[:10], fmt)
                        parsed_ok = True
                        break
                    except ValueError:
                        continue
                if not parsed_ok:
                    # openpyxl puede retornar un date/datetime object
                    import datetime
                    if not isinstance(raw, (datetime.date, datetime.datetime)):
                        row_errors.append(
                            f"El campo '{flabel}' debe ser una fecha (AAAA-MM-DD). Se encontró: '{val}'"
                        )

            # Tipo select
            elif ftype == "select":
                opts = _field_options(field_def)
                valid_values = set()
                for o in opts:
                    if isinstance(o, dict):
                        valid_values.add(str(o.get("value", "")).strip().lower())
                        valid_values.add(str(o.get("label", "")).strip().lower())
                    else:
                        valid_values.add(str(o).strip().lower())

                if valid_values and val.lower() not in valid_values:
                    opts_display = ", ".join(
                        str(o.get("value", o) if isinstance(o, dict) else o) for o in opts
                    )
                    row_errors.append(
                        f"El campo '{flabel}' tiene un valor inválido ('{val}'). "
                        f"Valores aceptados: {opts_display}"
                    )

        if row_errors:
            validation_errors.append({
                "fila": row_idx + data_start,  # número real en el Excel (1-indexed, + encabezados)
                "errores": row_errors,
            })

    if validation_errors:
        _uplog.warning("422 — %d fila(s) con errores de validación:", len(validation_errors))
        for ve in validation_errors:
            _uplog.warning("  Fila %s:", ve["fila"])
            for er in ve["errores"]:
                _uplog.warning("    · %s", er)
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "message": f"Se encontraron errores en {len(validation_errors)} fila(s). "
                           "Corrígelos y vuelve a cargar el archivo.",
                "errores_por_fila": validation_errors,
                "log_file": str(_upfile),
            },
        )

    # ── Crear formularios (todas las filas son válidas) ───────────────────
    lote_id = str(uuid.uuid4())
    created_ids: list[str] = []

    for row in data_rows:
        non_empty = [c for c in row if c is not None and str(c).strip() not in ("", "None")]
        if not non_empty:
            continue
        first_val = _str_val(row[0] if row else None).lower()
        if first_val in {"leyenda:", "nota:", "nota"} or first_val.startswith("leyenda"):
            continue

        datos_dinamicos: dict = {}
        informe_val = ""
        fecha_usuario: date | None = None

        # informe_cualitativo
        if informe_col is not None:
            informe_val = _str_val(_cell_val(row, informe_col))

        for col_idx, field_def in col_map.items():
            if not isinstance(field_def, dict):
                continue
            raw = _cell_val(row, col_idx)
            val = _str_val(raw)
            fname = _field_name(field_def)
            ftype = _field_type(field_def)
            readonly = _is_readonly(field_def)

            if fname == "informe_cualitativo":
                informe_val = val
                continue

            if fname == "mes_reporte" or fname == "fecha_referencia":
                if raw is not None:
                    import datetime
                    if isinstance(raw, (datetime.date, datetime.datetime)):
                        fecha_usuario = raw.date() if isinstance(raw, datetime.datetime) else raw
                    elif val:
                        from datetime import datetime as _dt
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                            try:
                                fecha_usuario = _dt.strptime(val[:10], fmt).date()
                                break
                            except ValueError:
                                continue
                continue

            if ftype == "computed":
                continue

            if not val and readonly and field_def.get("default") is not None:
                datos_dinamicos[fname] = field_def["default"]
            elif ftype == "number" and val:
                try:
                    datos_dinamicos[fname] = float(str(raw).replace(",", "."))
                except (TypeError, ValueError):
                    datos_dinamicos[fname] = None
            elif ftype == "date" and raw is not None:
                import datetime
                if isinstance(raw, (datetime.date, datetime.datetime)):
                    datos_dinamicos[fname] = raw.isoformat() if isinstance(raw, datetime.date) else raw.date().isoformat()
                else:
                    datos_dinamicos[fname] = val or None
            else:
                datos_dinamicos[fname] = val or None

        # Rellenar defaults de campos readonly no presentes en el Excel
        for f in all_fields:
            fname = _field_name(f)
            if _is_readonly(f) and f.get("default") is not None and not f.get("auto_calculate"):
                datos_dinamicos.setdefault(fname, f["default"])

        # Recalcular campos auto_calculate (sobreescribe cualquier valor del Excel)
        from app.services.auto_calc import recalc_auto_fields
        recalc_auto_fields(datos_dinamicos, all_fields)

        form = Form(
            plantilla_id=template_id,
            usuario_id=current_user.id,
            dependency_id=current_user.dependency_id,
            datos_dinamicos=datos_dinamicos,
            informe_cualitativo=informe_val or None,
            fecha_usuario=fecha_usuario or date.today(),
            estado=FormStatus.pending,
            cargado_via_excel=True,
            lote_excel_id=lote_id,
        )
        db.add(form)
        await db.flush()
        created_ids.append(str(form.id))

    if not created_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo no contiene filas de datos válidas.",
        )

    await db.commit()

    _uplog.info("═══ UPLOAD-EXCEL OK: %d formulario(s) creados, lote=%s ═══",
                len(created_ids), lote_id)
    return {
        "created": len(created_ids),
        "lote_id": lote_id,
        "form_ids": created_ids,
        "log_file": str(_upfile),
        "mensaje": f"Se cargaron {len(created_ids)} registro(s) correctamente y están en revisión. El validador revisará el lote completo.",
    }


# ── GET /forms/{form_id}/excel ────────────────────────────────────────────────

@router.get("/forms/{form_id}/excel")
async def download_form_as_excel(
    form_id: uuid.UUID,
    current_user: User = Depends(get_any_authenticated),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Descarga un formulario individual como archivo Excel."""
    openpyxl = _get_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    form_result = await db.execute(select(Form).where(Form.id == form_id))
    form = form_result.scalar_one_or_none()
    if form is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Formulario no encontrado")

    template_result = await db.execute(
        select(Template).where(Template.id == form.plantilla_id)
    )
    template = template_result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template no encontrado")

    all_fields = _get_fields(template)
    # Sólo exportar campos definidos en el template, excluyendo validator_only.
    export_cols = [f for f in all_fields if not f.get("validator_only") and not f.get("auto_calculate")]

    datos = form.datos_dinamicos or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (template.codigo or template.nombre)[:31]

    header_fill   = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    readonly_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    data_fill     = PatternFill(start_color="F8FDF8", end_color="F8FDF8", fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    data_font     = Font(size=10, name="Calibri")

    ws.append([_field_label(f) for f in export_cols])
    data_row = []
    for f in export_cols:
        fname = _field_name(f)
        ftype = _field_type(f)
        if fname == "informe_cualitativo":
            data_row.append(form.informe_cualitativo or "")
        elif fname in ("mes_reporte", "fecha_referencia"):
            data_row.append(str(form.fecha_usuario) if form.fecha_usuario else "")
        else:
            data_row.append(datos.get(fname, ""))
    ws.append(data_row)

    for col_idx, f in enumerate(export_cols, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 28
        h_cell = ws.cell(row=1, column=col_idx)
        h_cell.font = header_font
        h_cell.fill = header_fill
        h_cell.alignment = Alignment(horizontal="center", wrap_text=True)
        d_cell = ws.cell(row=2, column=col_idx)
        d_cell.font = data_font
        d_cell.fill = readonly_fill if _is_readonly(f) else data_fill
        d_cell.alignment = Alignment(horizontal="left", wrap_text=True)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"formulario_{form_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ════════════════════════════════════════════════════════════════════════════════
# Admin: exportar formularios de un template (Excel relleno + ZIP con evidencias)
# ════════════════════════════════════════════════════════════════════════════════

def _valid_estado_filter(estado: str | None) -> FormStatus | None:
    """Para DELETE y uso general. Admite None para 'todos'."""
    if estado in (None, "", "all", "todos"):
        return None
    try:
        return FormStatus(estado)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Estado inválido: '{estado}'. Válidos: draft, pending, approved, rejected o omitir.",
        )


def _cell_value_for_field(f: dict, form: Form) -> Any:
    """Extrae el valor a escribir en la celda Excel para un campo del template y un formulario."""
    fname = _field_name(f)
    ftype = _field_type(f)
    datos = form.datos_dinamicos or {}

    if fname == "informe_cualitativo":
        return form.informe_cualitativo or ""
    if fname in ("mes_reporte", "fecha_referencia"):
        return form.fecha_usuario.isoformat() if form.fecha_usuario else ""

    raw = datos.get(fname)
    if raw is None:
        return ""
    if ftype == "number":
        try:
            return float(raw)
        except (TypeError, ValueError):
            return raw
    if ftype == "date":
        # raw ya viene como ISO string desde el upload; se respeta
        return str(raw)[:10]
    return raw


def _build_forms_workbook(template: Template, forms: list[Form]):
    """
    Construye un openpyxl.Workbook con:
      Fila 1: Encabezados
      Fila 2: Tipo de dato/guía  (idéntica al Excel de ejemplo de carga)
      Fila 3+: una fila por formulario con los datos reales
    Devuelve (wb, export_cols).
    """
    openpyxl = _get_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fields = _get_fields(template)
    # Sólo exportar campos del template, sin validator_only.
    export_cols: list[dict] = [f for f in fields if not f.get("validator_only") and not f.get("auto_calculate")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (template.codigo or template.nombre)[:31]

    # ── Estilos (mismos que el Excel de ejemplo) ─────────────────────────
    header_fill    = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    type_row_fill  = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    header_font    = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    type_font      = Font(color="1A3C5E", italic=True, size=9, name="Calibri")
    data_font      = Font(size=10, name="Calibri")
    data_fill      = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    readonly_fill  = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    thin_border    = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Fila 1: Encabezados ──────────────────────────────────────────────
    headers = [_field_label(f) for f in export_cols]
    ws.append(headers)

    # ── Fila 2: Tipo de dato/guía ────────────────────────────────────────
    type_hints = []
    for f in export_cols:
        ft = _field_type(f)
        ro = _is_readonly(f)
        req = _is_required(f)
        if ro:
            hint = "Solo lectura"
        elif ft == "number":
            hint = "Número"
        elif ft == "date":
            hint = "Fecha (AAAA-MM-DD)"
        elif ft == "select":
            opts = _field_options(f)
            opts_str = " / ".join(str(o.get("value", o) if isinstance(o, dict) else o) for o in opts[:5])
            hint = f"Opciones: {opts_str}" if opts_str else "Seleccionar"
        elif ft == "textarea":
            hint = "Texto largo"
        else:
            hint = "Texto"
        if not ro:
            hint = ("* " if req else "") + hint
        type_hints.append(hint)
    ws.append(type_hints)

    # ── Filas 3+: datos de los formularios ──────────────────────────────
    for form in forms:
        row_data = [_cell_value_for_field(f, form) for f in export_cols]
        ws.append(row_data)

    # ── Estilos de fila 1 y 2 ───────────────────────────────────────────
    for col_idx, f in enumerate(export_cols, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 28

        h_cell = ws.cell(row=1, column=col_idx)
        h_cell.font = header_font
        h_cell.fill = header_fill
        h_cell.alignment = center_align
        h_cell.border = thin_border

        t_cell = ws.cell(row=2, column=col_idx)
        t_cell.font = type_font
        t_cell.fill = type_row_fill
        t_cell.alignment = left_align
        t_cell.border = thin_border

    # ── Estilos de celdas de datos ──────────────────────────────────────
    for row_idx in range(3, 3 + len(forms)):
        for col_idx, f in enumerate(export_cols, start=1):
            d_cell = ws.cell(row=row_idx, column=col_idx)
            d_cell.font = data_font
            d_cell.fill = readonly_fill if _is_readonly(f) else data_fill
            d_cell.alignment = left_align
            d_cell.border = thin_border

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28
    ws.sheet_view.zoomScale = 110

    # ── Hoja "Referencia" (informativa, NO contiene UUIDs) ────────────────
    # Permite que el admin vea metadatos básicos por fila, pero mantiene el
    # formato compatible con el upload normal de dependencia.
    ws_ref = wb.create_sheet("Referencia")
    ws_ref.append(["Fila Excel", "Estado", "Fecha", "Dependencia", "Carpeta ZIP sugerida"])
    for cell in ws_ref[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
        cell.alignment = center_align
    for idx, f in enumerate(forms, start=1):
        ws_ref.append([
            idx,
            f.estado.value if f.estado else "",
            f.fecha_usuario.isoformat() if f.fecha_usuario else "",
            f.dependency.nombre if f.dependency else "",
            f"formularios/{idx}/",
        ])
    for col in ["A", "B", "C", "D", "E"]:
        ws_ref.column_dimensions[col].width = 24

    return wb, export_cols


@router.get("/admin/templates/{template_id}/export-excel")
async def admin_export_template_forms_excel(
    template_id: uuid.UUID,
    current_user: User = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    [ADMIN] Descarga un Excel con todos los formularios APROBADOS de un template,
    en el MISMO formato que utilizan las dependencias al cargar (fila 1: headers,
    fila 2: tipo/guía, fila 3+: datos). Incluye hoja 'Metadatos' con correlación
    (ID, dependencia, usuario, lote).

    Solo se descargan registros con estado='approved' (única fuente confiable).
    """
    tpl_result = await db.execute(
        select(Template).where(Template.id == template_id)
    )
    template = tpl_result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template no encontrado")

    forms_q = (
        select(Form)
        .where(Form.plantilla_id == template_id, Form.estado == FormStatus.approved)
        .order_by(Form.fecha_carga.desc())
    )
    forms_result = await db.execute(forms_q)
    forms = list(forms_result.scalars().all())

    wb, _ = _build_forms_workbook(template, forms)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    slug = (template.codigo or template.nombre or "template").replace(" ", "_")[:50]
    filename = f"formularios_{slug}_aprobados.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/admin/templates/{template_id}/export-zip")
async def admin_export_template_forms_zip(
    template_id: uuid.UUID,
    current_user: User = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    [ADMIN] Descarga un ZIP con los formularios APROBADOS:
      - formularios_<slug>_aprobados.xlsx (mismo formato que el Excel de carga)
      - formularios/{form_id}/<archivo_original> …  (una carpeta por formulario)
      - _indice.csv con {form_id, fecha, dependencia, n_archivos}

    El ZIP sirve tanto como evidencia como para recargar en una dependencia
    (misma convención de nombres).
    """
    import csv
    import zipfile

    tpl_result = await db.execute(
        select(Template).where(Template.id == template_id)
    )
    template = tpl_result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template no encontrado")

    forms_q = (
        select(Form)
        .where(Form.plantilla_id == template_id, Form.estado == FormStatus.approved)
        .order_by(Form.fecha_carga.desc())
    )
    forms_result = await db.execute(forms_q)
    forms = list(forms_result.scalars().all())

    # Construir Excel en memoria
    wb, _ = _build_forms_workbook(template, forms)
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    # Construir ZIP en memoria
    zip_buf = io.BytesIO()
    minio = get_minio_service()
    slug = (template.codigo or template.nombre or "template").replace(" ", "_")[:50]

    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        # Excel consolidado en la raíz
        zf.writestr(f"formularios_{slug}_aprobados.xlsx", excel_buf.getvalue())

        # Índice CSV liviano (sin UUIDs) — correlación fila → carpeta ZIP
        idx_io = io.StringIO()
        writer = csv.writer(idx_io)
        writer.writerow([
            "fila", "estado", "fecha", "dependencia", "n_archivos", "carpeta_zip",
        ])

        # Una carpeta por formulario nombrada por su ÍNDICE DE FILA (1-based).
        # Esto coincide con el orden de filas en el Excel — así el ZIP se puede
        # recargar tal cual desde la vista de dependencia (misma convención que
        # el usuario usa cuando crea su propio ZIP manualmente).
        # Si un formulario NO tiene archivos, NO se crea la carpeta
        # (se SALTA ese número, tal como lo haría un usuario manualmente).
        for idx, f in enumerate(forms, start=1):
            folder = f"formularios/{idx}"
            n_archivos = 0
            for arch in (f.archivos or []):
                try:
                    resp = minio.get_file_stream(arch)
                    try:
                        data = resp.read()
                    finally:
                        try:
                            resp.close()
                            resp.release_conn()
                        except Exception:
                            pass
                    base_name = arch.nombre_original or arch.nombre_minio
                    target = f"{folder}/{base_name}"
                    safe_target = target.replace("..", "_").replace("\\", "/")
                    try:
                        zf.writestr(safe_target, data)
                        n_archivos += 1
                    except Exception as exc:
                        logger.warning("No se pudo escribir archivo %s al ZIP: %s", arch.id, exc)
                except Exception as exc:
                    logger.warning("No se pudo leer archivo %s desde MinIO: %s", arch.id, exc)

            writer.writerow([
                idx,
                f.estado.value if f.estado else "",
                f.fecha_usuario.isoformat() if f.fecha_usuario else "",
                f.dependency.nombre if f.dependency else "",
                n_archivos,
                folder if n_archivos > 0 else "(sin adjuntos)",
            ])

        zf.writestr("_indice.csv", idx_io.getvalue())

    zip_buf.seek(0)
    filename = f"formularios_{slug}_aprobados.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ════════════════════════════════════════════════════════════════════════════════
# Admin: eliminar registros (y template) — habilita el ciclo descargar → borrar
# → recargar descrito en la UI
# ════════════════════════════════════════════════════════════════════════════════

@router.get("/admin/templates/counts")
async def admin_template_forms_counts(
    current_user: User = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
) -> dict[str, dict[str, int]]:
    """
    [ADMIN] Retorna el conteo de formularios por template y por estado.
    Útil para habilitar/deshabilitar acciones en la UI ("Borrar template"
    solo si no hay registros, "Descargar aprobados" solo si hay aprobados).

    Formato: {
        "<template_uuid>": {"total": int, "approved": int, "pending": int,
                            "draft": int, "rejected": int}
    }
    """
    from sqlalchemy import func
    result = await db.execute(
        select(Form.plantilla_id, Form.estado, func.count(Form.id))
        .group_by(Form.plantilla_id, Form.estado)
    )
    counts: dict[str, dict[str, int]] = {}
    for plantilla_id, estado, n in result.all():
        tid = str(plantilla_id)
        row = counts.setdefault(tid, {"total": 0, "approved": 0, "pending": 0, "draft": 0, "rejected": 0})
        row["total"] += n
        if estado:
            key = estado.value if hasattr(estado, "value") else str(estado)
            if key in row:
                row[key] += n
    return counts


@router.delete("/admin/templates/{template_id}/registros")
async def admin_delete_template_forms(
    template_id: uuid.UUID,
    estado: str | None = None,
    current_user: User = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    [ADMIN] Elimina formularios respondidos de un template (y sus archivos MinIO).

    Query:
        estado=approved|pending|rejected|draft → borra solo los de ese estado
        (si se omite, borra TODOS los registros del template).
    """
    from sqlalchemy import delete as sa_delete

    estado_enum = _valid_estado_filter(estado)

    tpl_result = await db.execute(
        select(Template).where(Template.id == template_id)
    )
    template = tpl_result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template no encontrado")

    forms_q = select(Form).where(Form.plantilla_id == template_id)
    if estado_enum is not None:
        forms_q = forms_q.where(Form.estado == estado_enum)
    forms = list((await db.execute(forms_q)).scalars().all())

    # Eliminar archivos en MinIO antes de borrar el registro en BD
    minio = get_minio_service()
    n_archivos = 0
    for f in forms:
        for arch in (f.archivos or []):
            try:
                minio.delete_file(arch)
                n_archivos += 1
            except Exception as exc:
                logger.warning("No se pudo eliminar archivo %s en MinIO: %s", arch.id, exc)

    # Borrar en cascada (archivos en BD se eliminan por cascade="all, delete-orphan")
    ids_to_delete = [f.id for f in forms]
    if ids_to_delete:
        await db.execute(sa_delete(Form).where(Form.id.in_(ids_to_delete)))
        await db.commit()

    return {
        "ok": True,
        "registros_eliminados": len(ids_to_delete),
        "archivos_eliminados": n_archivos,
        "estado_filtro": estado_enum.value if estado_enum else "todos",
        "mensaje": f"Se eliminaron {len(ids_to_delete)} registro(s) y {n_archivos} archivo(s).",
    }


@router.delete("/admin/templates/{template_id}")
async def admin_delete_template(
    template_id: uuid.UUID,
    current_user: User = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    [ADMIN] Elimina un template. Requiere que NO tenga formularios asociados
    (en cualquier estado). Si los tiene, responde 409 y pide vaciarlos primero.
    """
    from sqlalchemy import delete as sa_delete, func

    tpl_result = await db.execute(
        select(Template).where(Template.id == template_id)
    )
    template = tpl_result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template no encontrado")

    n_forms = await db.scalar(
        select(func.count(Form.id)).where(Form.plantilla_id == template_id)
    ) or 0
    if n_forms > 0:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"El template aún tiene {n_forms} formulario(s). Elimina los registros antes de borrar el template.",
        )

    nombre = template.nombre
    await db.execute(sa_delete(Template).where(Template.id == template_id))
    await db.commit()

    return {
        "ok": True,
        "mensaje": f"Template '{nombre}' eliminado.",
    }


# ════════════════════════════════════════════════════════════════════════════════
# Dependencia: upload Excel + ZIP con adjuntos (ciclo de recarga)
# ════════════════════════════════════════════════════════════════════════════════

def _parse_zip_folder_mapping(
    zip_bytes: bytes,
    uuid_by_row: dict[int, str] | None,
    max_rows: int,
) -> tuple[dict[int, list[tuple[str, bytes]]], list[str]]:
    """
    Lee un ZIP y retorna (mapping, errors).

    mapping: { fila_index (1..max_rows): [(nombre, bytes), ...] }
    errors:  lista de mensajes legibles si la estructura no cumple la convención.

    Acepta:
        formularios/1/archivo.pdf
        formularios/fila_1/archivo.pdf
        formularios/{uuid}/archivo.pdf   (si hay hoja Metadatos antigua)
        1/archivo.pdf                    (sin subcarpeta)
        fila_1/archivo.pdf

    Rechaza:
        - archivos en la raíz del zip
        - carpetas con identificador no numérico ni UUID conocido
        - números de fila fuera de rango (ej: formularios/10 cuando solo hay 3 filas)
    """
    import zipfile
    mapping: dict[int, list[tuple[str, bytes]]] = {}
    errors: list[str] = []

    if not zip_bytes:
        return mapping, errors

    uuid_to_row: dict[str, int] = {}
    if uuid_by_row:
        uuid_to_row = {v.lower(): k for k, v in uuid_by_row.items()}

    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        errors.append("El archivo .zip no es válido o está corrupto.")
        return mapping, errors

    carpetas_invalidas: set[str] = set()
    archivos_en_raiz: list[str] = []
    carpetas_fuera_rango: dict[str, int] = {}

    for info in zf.infolist():
        if info.is_dir():
            continue

        path = info.filename.replace("\\", "/").lstrip("./")
        # Ignorar metadatos de Mac/Windows y archivos auxiliares del propio export admin
        low_path = path.lower()
        if (low_path.startswith("__macosx/")
                or low_path.startswith("._")
                or "/__macosx/" in low_path
                or low_path.endswith(".ds_store")
                or low_path.endswith("/_indice.csv")
                or low_path == "_indice.csv"
                or low_path.endswith(".xlsx") or low_path.endswith(".xls")):
            continue

        parts = [p for p in path.split("/") if p and not p.startswith(".")]
        if len(parts) < 2:
            archivos_en_raiz.append(path)
            continue

        # Detectar estructura
        if parts[0].lower() in {"formularios", "forms", "registros"} and len(parts) >= 3:
            ident, filename = parts[1], parts[-1]
        else:
            ident, filename = parts[0], parts[-1]

        row_idx: int | None = None
        ident_low = ident.lower().strip()

        if ident_low.isdigit():
            row_idx = int(ident_low)
        else:
            for prefix in ("fila_", "formulario_", "row_", "f_"):
                if ident_low.startswith(prefix):
                    rem = ident_low[len(prefix):]
                    if rem.isdigit():
                        row_idx = int(rem)
                        break

        if row_idx is None and ident_low in uuid_to_row:
            row_idx = uuid_to_row[ident_low]
        if row_idx is None:
            try:
                parsed = str(uuid.UUID(ident_low)).lower()
                if parsed in uuid_to_row:
                    row_idx = uuid_to_row[parsed]
            except (ValueError, AttributeError):
                pass

        if row_idx is None:
            carpetas_invalidas.add(ident)
            continue

        if row_idx < 1 or row_idx > max_rows:
            carpetas_fuera_rango[ident] = row_idx
            continue

        try:
            data = zf.read(info.filename)
        except Exception as exc:
            errors.append(f"No se pudo leer '{path}' del ZIP: {exc}")
            continue

        mapping.setdefault(row_idx, []).append((filename, data))

    if archivos_en_raiz:
        sample = ", ".join(archivos_en_raiz[:3]) + ("…" if len(archivos_en_raiz) > 3 else "")
        errors.append(
            f"El ZIP tiene {len(archivos_en_raiz)} archivo(s) en la raíz "
            f"({sample}). Debes ponerlos dentro de carpetas 'formularios/N/' donde N es el número de fila."
        )
    if carpetas_invalidas:
        sample = ", ".join(sorted(carpetas_invalidas)[:5])
        errors.append(
            f"Carpetas con nombre inválido en el ZIP: [{sample}]. "
            f"Usa números (1, 2, 3…) correspondientes a la fila del Excel."
        )
    if carpetas_fuera_rango:
        detalle = ", ".join(f"'{k}' (fila {v})" for k, v in list(carpetas_fuera_rango.items())[:5])
        errors.append(
            f"Estas carpetas del ZIP no corresponden a ninguna fila del Excel "
            f"({max_rows} fila(s) de datos): {detalle}."
        )

    return mapping, errors


def _extract_uuid_metadata_from_excel(wb) -> dict[int, str]:
    """
    Backward-compat: si el Excel tiene una hoja 'Metadatos' antigua con una
    columna de UUID en la primera posición, retorna un dict {fila_index: uuid}.
    Devuelve vacío si no hay tal hoja o si la primera columna no parece un UUID.

    (Los exports nuevos usan la hoja 'Referencia' con índices de fila y no
    necesitan este mapeo — basta con el nombre numérico de la carpeta.)
    """
    mapping: dict[int, str] = {}
    if "Metadatos" not in wb.sheetnames:
        return mapping
    ws = wb["Metadatos"]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row or not row[0]:
            continue
        val = str(row[0]).strip()
        # solo usar si parece un UUID
        try:
            uuid.UUID(val)
        except (ValueError, AttributeError):
            continue
        mapping[i] = val
    return mapping


@router.post("/forms/upload-excel-zip/{template_id}", status_code=status.HTTP_201_CREATED)
async def upload_excel_with_attachments(
    template_id: uuid.UUID,
    file: UploadFile = File(..., description="Archivo Excel con los datos"),
    archivos_zip: UploadFile | None = File(None, description="ZIP opcional con evidencias"),
    current_user: User = Depends(get_dependency_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    [DEPENDENCIA] Carga un Excel (mismo formato que el ejemplo descargable) junto
    con un ZIP opcional que contiene las evidencias adjuntas agrupadas por
    carpetas identificadoras.

    Convenciones aceptadas para los nombres de carpeta dentro del ZIP:

      formularios/1/<archivo>              (fila 1 de datos)
      formularios/2/<archivo>              (fila 2 de datos)
      formularios/{uuid}/<archivo>         (si el Excel trae hoja 'Metadatos')
      1/<archivo>                          (sin subcarpeta 'formularios')
      fila_1/<archivo>                     (alternativa textual)

    El ZIP puede generarse manualmente o ser el mismo que exportó el admin
    (en cuyo caso las carpetas se llaman {uuid} y se mapean automáticamente
    al orden de los registros en el Excel).
    """
    openpyxl = _get_openpyxl()

    # 1) Cargar template
    t_result = await db.execute(
        select(Template).where(Template.id == template_id, Template.activo == True)
    )
    template = t_result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template no encontrado")

    all_fields = _get_fields(template)
    field_by_label = {_field_label(f): f for f in all_fields}
    field_by_name  = {_field_name(f):  f for f in all_fields}
    INFORME_LABEL = "Informe cualitativo de la búsqueda"
    template_has_informe_field = any(_field_name(f) == "informe_cualitativo" for f in all_fields)

    # 2) Leer Excel
    excel_bytes = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se pudo leer el archivo Excel: {e}",
        )

    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El Excel debe tener encabezado y al menos una fila de datos.",
        )

    header_row = all_rows[0]
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    data_start = 1
    if len(all_rows) >= 2:
        row2_values = [str(v).strip().lower() if v is not None else "" for v in all_rows[1]]
        type_hint_keywords = {"número", "texto", "fecha", "opciones", "solo lectura", "texto largo", "seleccionar"}
        if any(any(kw in v for kw in type_hint_keywords) for v in row2_values if v):
            data_start = 2
    data_rows = all_rows[data_start:]

    # Mapear columnas → campo del template
    col_map: dict[int, dict] = {}
    informe_col: int | None = None
    for col_idx, header in enumerate(headers):
        h = header.strip()
        if not h:
            continue
        mapped_f = field_by_label.get(h) or field_by_name.get(h)
        is_informe = (
            h == INFORME_LABEL or h.lower() == "informe_cualitativo"
            or (mapped_f is not None and _field_name(mapped_f) == "informe_cualitativo")
        )
        if is_informe:
            informe_col = col_idx
            continue
        # Ignorar columnas validator_only / auto_calculate (no las llena la dependencia)
        if mapped_f is not None and (mapped_f.get("validator_only") or mapped_f.get("auto_calculate")):
            continue
        # Columnas no reconocidas → IGNORAR (no error)
        if h in field_by_label:
            col_map[col_idx] = field_by_label[h]
        elif h in field_by_name:
            col_map[col_idx] = field_by_name[h]

    # 3) Leer ZIP si viene
    zip_bytes = b""
    if archivos_zip is not None:
        # validar que sea zip por extensión al menos
        fname = (archivos_zip.filename or "").lower()
        if fname and not fname.endswith(".zip"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "message": "El archivo de adjuntos debe ser un .zip",
                    "excel_errors": [],
                    "zip_errors": [f"El archivo '{archivos_zip.filename}' no es un .zip"],
                    "estructura_errors": [],
                },
            )
        zip_bytes = await archivos_zip.read()

    # 4) Validar filas del Excel (antes de tocar MinIO o la BD)
    def _str_val(v: Any) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        return "" if s.lower() == "none" else s

    def _cell_val(row: tuple, idx: int) -> Any:
        return row[idx] if idx < len(row) else None

    validation_errors: list[dict] = []
    estructura_errors: list[str] = []
    real_row_positions: list[int] = []

    # Validación estructural: debe haber al menos un campo del template mapeado
    if not col_map and informe_col is None:
        estructura_errors.append(
            "El Excel no tiene ninguna columna que coincida con los campos del template. "
            "Descarga el Excel de ejemplo y úsalo como base."
        )

    for row_idx, row in enumerate(data_rows):
        non_empty = [c for c in row if c is not None and str(c).strip() not in ("", "None")]
        if not non_empty:
            continue
        first_val = _str_val(row[0] if row else None).lower()
        if first_val.startswith("leyenda") or first_val in {"nota:", "nota"}:
            continue
        real_row_positions.append(row_idx)

        row_errors: list[str] = []
        # informe_cualitativo es opcional (eliminado del flujo del PAI 2026)
        informe_val = _str_val(_cell_val(row, informe_col)) if informe_col is not None else ""

        for col_idx, field_def in col_map.items():
            raw = _cell_val(row, col_idx)
            val = _str_val(raw)
            ftype = _field_type(field_def)
            flabel = _field_label(field_def)
            # No validar readonly, validator_only ni auto_calculate
            if (_is_readonly(field_def)
                or field_def.get("validator_only")
                or field_def.get("auto_calculate")):
                continue
            if _is_required(field_def) and not val:
                row_errors.append(f"El campo '{flabel}' es obligatorio")
                continue
            if not val:
                continue
            if ftype == "number":
                try:
                    float(str(raw).replace(",", "."))
                except (TypeError, ValueError):
                    row_errors.append(f"El campo '{flabel}' debe ser un número (se encontró: '{val}')")
            elif ftype == "date":
                from datetime import datetime as _dt
                parsed_ok = False
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                    try:
                        _dt.strptime(val[:10], fmt)
                        parsed_ok = True
                        break
                    except ValueError:
                        continue
                if not parsed_ok:
                    import datetime as _dtmod
                    if not isinstance(raw, (_dtmod.date, _dtmod.datetime)):
                        row_errors.append(
                            f"El campo '{flabel}' debe ser una fecha (AAAA-MM-DD). Se encontró: '{val}'"
                        )
            elif ftype == "select":
                opts = _field_options(field_def)
                valid_values = set()
                for o in opts:
                    if isinstance(o, dict):
                        valid_values.add(str(o.get("value", "")).strip().lower())
                        valid_values.add(str(o.get("label", "")).strip().lower())
                    else:
                        valid_values.add(str(o).strip().lower())
                if valid_values and val.lower() not in valid_values:
                    opts_display = ", ".join(
                        str(o.get("value", o) if isinstance(o, dict) else o) for o in opts
                    )
                    row_errors.append(
                        f"El campo '{flabel}' tiene un valor inválido ('{val}'). Valores aceptados: {opts_display}"
                    )
        if row_errors:
            validation_errors.append({
                "fila": row_idx + data_start + 1,
                "errores": row_errors,
            })

    if not real_row_positions:
        estructura_errors.append("El Excel no contiene filas de datos válidas.")

    # 5) Validar estructura del ZIP (si vino) contra el Excel
    uuid_by_row = _extract_uuid_metadata_from_excel(wb)
    attachments_by_row, zip_errors = _parse_zip_folder_mapping(
        zip_bytes, uuid_by_row, max_rows=len(real_row_positions),
    )

    # 6) Si hay CUALQUIER error, rechazar todo — no se crea nada
    if validation_errors or zip_errors or estructura_errors:
        total = len(validation_errors) + len(zip_errors) + len(estructura_errors)
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "message": (
                    f"Se encontraron {total} problema(s) de validación. "
                    "La carga fue RECHAZADA y no se creó ningún registro. "
                    "Corrige los errores y vuelve a intentarlo."
                ),
                "excel_errors": validation_errors,
                "zip_errors": zip_errors,
                "estructura_errors": estructura_errors,
            },
        )

    # 7) Crear formularios y adjuntar archivos (ATÓMICO — rollback MinIO si falla)
    from datetime import date as _date, datetime as _dt
    minio = get_minio_service()
    lote_id = str(uuid.uuid4())
    created_ids: list[str] = []
    total_archivos = 0
    uploaded_minio_paths: list[tuple[str, str]] = []  # para rollback

    try:
        for position, data_row_idx in enumerate(real_row_positions, start=1):
            row = data_rows[data_row_idx]
            datos_dinamicos: dict = {}
            informe_val = ""
            fecha_usuario: _date | None = None

            if informe_col is not None:
                informe_val = _str_val(_cell_val(row, informe_col))

            for col_idx, field_def in col_map.items():
                raw = _cell_val(row, col_idx)
                val = _str_val(raw)
                fname = _field_name(field_def)
                ftype = _field_type(field_def)
                readonly = _is_readonly(field_def)
                if fname == "informe_cualitativo":
                    informe_val = val
                    continue
                if fname in ("mes_reporte", "fecha_referencia"):
                    if raw is not None:
                        import datetime as _dtmod
                        if isinstance(raw, (_dtmod.date, _dtmod.datetime)):
                            fecha_usuario = raw.date() if isinstance(raw, _dtmod.datetime) else raw
                        elif val:
                            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                                try:
                                    fecha_usuario = _dt.strptime(val[:10], fmt).date()
                                    break
                                except ValueError:
                                    continue
                    continue
                if ftype == "computed":
                    continue
                if not val and readonly and field_def.get("default") is not None:
                    datos_dinamicos[fname] = field_def["default"]
                elif ftype == "number" and val:
                    try:
                        datos_dinamicos[fname] = float(str(raw).replace(",", "."))
                    except (TypeError, ValueError):
                        datos_dinamicos[fname] = None
                elif ftype == "date" and raw is not None:
                    import datetime as _dtmod
                    if isinstance(raw, (_dtmod.date, _dtmod.datetime)):
                        datos_dinamicos[fname] = raw.isoformat() if isinstance(raw, _dtmod.date) else raw.date().isoformat()
                    else:
                        datos_dinamicos[fname] = val or None
                else:
                    datos_dinamicos[fname] = val or None

            for f in all_fields:
                fname = _field_name(f)
                if _is_readonly(f) and f.get("default") is not None and not f.get("auto_calculate"):
                    datos_dinamicos.setdefault(fname, f["default"])

            # Recalcular campos auto_calculate (sobreescribe el valor del Excel)
            from app.services.auto_calc import recalc_auto_fields
            recalc_auto_fields(datos_dinamicos, all_fields)

            form = Form(
                plantilla_id=template_id,
                usuario_id=current_user.id,
                dependency_id=current_user.dependency_id,
                datos_dinamicos=datos_dinamicos,
                informe_cualitativo=informe_val or None,
                fecha_usuario=fecha_usuario or _date.today(),
                estado=FormStatus.pending,
                cargado_via_excel=True,
                lote_excel_id=lote_id,
            )
            db.add(form)
            await db.flush()
            created_ids.append(str(form.id))

            # Adjuntar archivos del ZIP correspondientes a esta fila
            row_files = attachments_by_row.get(position, [])
            for original_name, content in row_files:
                ext = ""
                if "." in original_name:
                    ext = "." + original_name.rsplit(".", 1)[-1]
                unique_name = f"{uuid.uuid4()}{ext}"
                object_path = minio._get_object_path(current_user.dependency_id, form.id, unique_name)
                import mimetypes
                mime = mimetypes.guess_type(original_name)[0] or "application/octet-stream"

                minio.client.put_object(
                    bucket_name=minio.bucket,
                    object_name=object_path,
                    data=io.BytesIO(content),
                    length=len(content),
                    content_type=mime,
                )
                uploaded_minio_paths.append((minio.bucket, object_path))

                arch = Archivo(
                    formulario_id=form.id,
                    nombre_original=original_name,
                    nombre_minio=unique_name,
                    bucket=minio.bucket,
                    ruta_minio=object_path,
                    tipo_mime=mime,
                    tamaño_bytes=len(content),
                )
                db.add(arch)
                total_archivos += 1

        await db.commit()
    except Exception as exc:
        # Rollback BD (no se hizo commit aún) + limpieza MinIO
        await db.rollback()
        for b, p in uploaded_minio_paths:
            try:
                minio.client.remove_object(b, p)
            except Exception as rb_exc:
                logger.warning("Rollback MinIO: no se pudo borrar %s: %s", p, rb_exc)
        if isinstance(exc, HTTPException):
            raise
        logger.exception("Error al crear formularios desde Excel+ZIP")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "message": "Error al procesar la carga. Ninguna fila fue creada.",
                "excel_errors": [],
                "zip_errors": [],
                "estructura_errors": [str(exc)],
            },
        )

    return {
        "created": len(created_ids),
        "archivos_subidos": total_archivos,
        "lote_id": lote_id,
        "form_ids": created_ids,
        "mensaje": (
            f"Se cargaron {len(created_ids)} registro(s) "
            + (f"con {total_archivos} archivo(s) adjunto(s)" if total_archivos else "(sin adjuntos)")
            + ". El validador revisará el lote completo."
        ),
    }
